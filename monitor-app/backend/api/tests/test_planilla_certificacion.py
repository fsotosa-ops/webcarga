"""La planilla de certificación, ejercitada contra Postgres de verdad.

POR QUÉ ACÁ Y NO CON AsyncMock. Un mock devuelve las claves que se le dicen que
devuelva, así que nunca contradice al SQL. En este mismo router eso ya dejó
pasar un 500 con las 916 y las 1.292 en verde (renombrar un documento, 21/08).
La planilla escribe con `unnest` sobre arreglos y lee con un `ANY($1::uuid[])`:
si alguno de esos casts está mal, un mock no se entera.

CÓMO SE PROTEGE LA BASE. Es la única base que hay, o sea producción. Cada test
recibe `conexion_revertida`, una conexión ya adentro de una transacción que el
fixture revierte en un `finally`, y los datos los crea el propio test ahí
adentro. Además `guardia_de_produccion` toma la huella de
public.compliance_records antes y después de la sesión.

    venv/bin/python -m pytest tests/test_planilla_certificacion.py -q
"""
from __future__ import annotations

import io
import re
from datetime import date, timedelta
from uuid import uuid4

import pytest
from fastapi import HTTPException, UploadFile

from app.routers.compliance import (
    _filas_de_planilla,
    _planilla_a_xlsx,
    cargar_planilla,
    resumen_de_planilla,
)
from app.services.plantilla_certificacion import (
    COLUMNAS,
    COLUMNA_LLAVE,
    COLUMNA_TENENCIA,
    COLUMNA_VENCIMIENTO,
    FUENTE_AUDITORIA,
    SQL_APLICAR,
    SQL_AUDITAR,
)
from tests.conftest import PoolDeUnaConexion, _usuario_real

pytestmark = pytest.mark.integracion

PREFIJO = "ZZ-TEST-PLANILLA"


# ── Andamiaje ────────────────────────────────────────────────────────────────


async def _empresa_activa(conn) -> str:
    """Una empresa ACTIVE. Al insertarla, `trg_reconcile_new_carrier` le siembra
    sus compliance_records en MISSING — o sea que las filas de la planilla las
    produce el trigger real, no un INSERT a mano que podría no parecerse."""
    suf = uuid4().hex[:10]
    return await conn.fetchval(
        "INSERT INTO public.carriers (business_name, tax_id, operational_status) "
        "VALUES ($1, $2, 'ACTIVE') RETURNING id",
        f"{PREFIJO} {suf}", f"{PREFIJO}-{suf}",
    )


async def _registros_de(conn, empresa, *, con_vencimiento: bool) -> list:
    return await conn.fetch(
        """
        SELECT cr.id::text AS id, req.name AS documento, cr.status, cr.expiration_date
        FROM public.compliance_records cr
        JOIN public.compliance_requirements req ON req.id = cr.requirement_id
        WHERE cr.entity_type = 'CARRIER' AND cr.entity_id = $1
          AND cr.is_current = true AND req.is_active = true
          AND req.has_expiration = $2
        ORDER BY req.name
        """,
        empresa, con_vencimiento,
    )


def _subir(filas: list[dict], *, nombre: str = "certificacion.xlsx") -> UploadFile:
    """Escribe con el MISMO escritor que baja la planilla. Es el punto: el test
    no inventa un formato, usa el que la aplicación produce."""
    return UploadFile(file=io.BytesIO(_planilla_a_xlsx(filas)), filename=nombre)


async def _fila(pool, registro_id: str) -> dict:
    for fila in await _filas_de_planilla(pool, "todas"):
        if fila[COLUMNA_LLAVE] == registro_id:
            return fila
    raise AssertionError(f"la planilla no trajo el registro {registro_id}")


# ── Los dos ejes ─────────────────────────────────────────────────────────────


async def test_la_planilla_trae_tambien_los_documentos_que_no_vencen(conexion_revertida):
    """En las 39 empresas activas, 1.044 de los 2.370 pendientes NO llevan
    vencimiento y son TODOS obligatorios. Dejarlos fuera deja a la mitad del
    trabajo pendiente sin ninguna vía de carga: su pregunta es la tenencia."""
    pool = PoolDeUnaConexion(conexion_revertida)
    empresa = await _empresa_activa(conexion_revertida)

    con_fecha = await _registros_de(conexion_revertida, empresa, con_vencimiento=True)
    sin_fecha = await _registros_de(conexion_revertida, empresa, con_vencimiento=False)
    assert con_fecha and sin_fecha, "la empresa sintética tiene que tener de los dos"

    en_planilla = {f[COLUMNA_LLAVE]: f for f in await _filas_de_planilla(pool, "activas")}

    assert {r["id"] for r in con_fecha} <= set(en_planilla)
    assert {r["id"] for r in sin_fecha} <= set(en_planilla), (
        "quedaron fuera documentos obligatorios que no vencen"
    )
    # Y llegan con la celda de fecha vacía: no hay fecha que declarar.
    for registro in sin_fecha:
        assert en_planilla[registro["id"]][COLUMNA_VENCIMIENTO] == ""


async def test_declarar_tenencia_en_un_documento_que_no_vence(conexion_revertida):
    """El caso de los 1.044: la única pregunta posible es si lo tenemos."""
    pool = PoolDeUnaConexion(conexion_revertida)
    usuario = await _usuario_real(conexion_revertida)
    empresa = await _empresa_activa(conexion_revertida)
    objetivo = (await _registros_de(conexion_revertida, empresa, con_vencimiento=False))[0]

    fila = await _fila(pool, objetivo["id"])
    fila[COLUMNA_TENENCIA] = "Sí"

    resultado = await cargar_planilla(file=_subir([fila]), dry_run=False, pool=pool, user=usuario)
    assert resultado["cambian"] == 1 and resultado["recibidos"] == 1
    assert resultado["total_errores"] == 0, resultado["errores"]

    guardado = await conexion_revertida.fetchrow(
        "SELECT status, expiration_date, is_manual_override "
        "FROM public.compliance_records WHERE id = $1::uuid", objetivo["id"],
    )
    assert guardado["status"] == "APPROVED_MANUAL"
    assert guardado["expiration_date"] is None
    assert guardado["is_manual_override"] is True


async def test_no_se_da_por_recibido_lo_que_vence_sin_su_fecha(conexion_revertida):
    """Es el defecto que dejó 14 documentos invisibles: aprobados con
    expiration_date NULL desaparecen de pendientes para siempre, aunque el papel
    real venza el mes que viene."""
    pool = PoolDeUnaConexion(conexion_revertida)
    usuario = await _usuario_real(conexion_revertida)
    empresa = await _empresa_activa(conexion_revertida)
    objetivo = (await _registros_de(conexion_revertida, empresa, con_vencimiento=True))[0]

    fila = await _fila(pool, objetivo["id"])
    fila[COLUMNA_TENENCIA] = "Sí"          # sin fecha
    fila[COLUMNA_VENCIMIENTO] = ""

    with pytest.raises(HTTPException) as caso:
        await cargar_planilla(file=_subir([fila]), dry_run=False, pool=pool, user=usuario)
    assert caso.value.status_code == 422
    assert "necesita su fecha de vencimiento" in str(caso.value.detail)


async def test_la_fecha_sola_no_aprueba_el_documento(conexion_revertida):
    """Los dos ejes no se colapsan: mover la vigencia no puede tocar la
    evidencia. Si el status pasara a aprobado, cargar miles de fechas pondría
    empresas en verde sin que nadie haya visto un papel."""
    pool = PoolDeUnaConexion(conexion_revertida)
    usuario = await _usuario_real(conexion_revertida)
    empresa = await _empresa_activa(conexion_revertida)
    objetivo = (await _registros_de(conexion_revertida, empresa, con_vencimiento=True))[0]

    fila = await _fila(pool, objetivo["id"])
    fila[COLUMNA_TENENCIA] = ""
    fila[COLUMNA_VENCIMIENTO] = (date.today() + timedelta(days=60)).strftime("%d-%m-%Y")
    await cargar_planilla(file=_subir([fila]), dry_run=False, pool=pool, user=usuario)

    guardado = await conexion_revertida.fetchrow(
        "SELECT status, expiration_date, file_url FROM public.compliance_records WHERE id = $1::uuid",
        objetivo["id"],
    )
    assert guardado["status"] == "MISSING"
    assert guardado["file_url"] is None
    assert guardado["expiration_date"] == date.today() + timedelta(days=60)


async def test_no_acepta_fecha_en_un_documento_que_no_vence(conexion_revertida):
    pool = PoolDeUnaConexion(conexion_revertida)
    usuario = await _usuario_real(conexion_revertida)
    empresa = await _empresa_activa(conexion_revertida)
    objetivo = (await _registros_de(conexion_revertida, empresa, con_vencimiento=False))[0]

    fila = await _fila(pool, objetivo["id"])
    fila[COLUMNA_VENCIMIENTO] = "01-01-2027"

    with pytest.raises(HTTPException) as caso:
        await cargar_planilla(file=_subir([fila]), dry_run=False, pool=pool, user=usuario)
    assert "no lleva fecha de vencimiento" in str(caso.value.detail)


# ── El ida y vuelta ──────────────────────────────────────────────────────────


async def test_una_planilla_parcial_no_toca_el_resto(conexion_revertida):
    """Pedido explícito: puede pasar que sólo se actualicen los conductores y no
    se cargue todo. Subir un subconjunto de filas tiene que ser válido."""
    pool = PoolDeUnaConexion(conexion_revertida)
    usuario = await _usuario_real(conexion_revertida)
    empresa = await _empresa_activa(conexion_revertida)
    registros = await _registros_de(conexion_revertida, empresa, con_vencimiento=False)
    tocado, intacto = registros[0], registros[1]

    fila = await _fila(pool, tocado["id"])
    fila[COLUMNA_TENENCIA] = "Sí"

    # UNA sola fila de las 13 que tiene la empresa, y de las miles de la planilla
    resultado = await cargar_planilla(file=_subir([fila]), dry_run=False, pool=pool, user=usuario)
    assert resultado["cambian"] == 1 and resultado["vacias"] == 0

    quedo = await conexion_revertida.fetchval(
        "SELECT status FROM public.compliance_records WHERE id = $1::uuid", intacto["id"])
    assert quedo == "MISSING", "una planilla parcial movió una fila que no venía en ella"


async def test_la_fila_vacia_no_se_toca(conexion_revertida):
    """Es lo que permite bajar la planilla entera y llenar sólo lo que se sabe.
    Vaciar una celda es "no sé", no "no vence"."""
    pool = PoolDeUnaConexion(conexion_revertida)
    usuario = await _usuario_real(conexion_revertida)
    empresa = await _empresa_activa(conexion_revertida)
    registros = await _registros_de(conexion_revertida, empresa, con_vencimiento=True)
    ya_tiene, a_llenar = registros[0], registros[1]
    previa = date.today() + timedelta(days=200)
    await conexion_revertida.execute(
        "UPDATE public.compliance_records SET expiration_date = $2 WHERE id = $1::uuid",
        ya_tiene["id"], previa,
    )

    vacia = await _fila(pool, ya_tiene["id"])
    vacia[COLUMNA_TENENCIA] = ""
    vacia[COLUMNA_VENCIMIENTO] = ""
    llena = await _fila(pool, a_llenar["id"])
    llena[COLUMNA_TENENCIA] = ""
    llena[COLUMNA_VENCIMIENTO] = (date.today() + timedelta(days=30)).strftime("%d-%m-%Y")

    resultado = await cargar_planilla(
        file=_subir([vacia, llena]), dry_run=False, pool=pool, user=usuario)

    assert resultado["vacias"] == 1 and resultado["cambian"] == 1
    assert await conexion_revertida.fetchval(
        "SELECT expiration_date FROM public.compliance_records WHERE id = $1::uuid",
        ya_tiene["id"]) == previa


async def test_dry_run_cuenta_y_no_escribe(conexion_revertida):
    pool = PoolDeUnaConexion(conexion_revertida)
    usuario = await _usuario_real(conexion_revertida)
    empresa = await _empresa_activa(conexion_revertida)
    objetivo = (await _registros_de(conexion_revertida, empresa, con_vencimiento=False))[0]

    fila = await _fila(pool, objetivo["id"])
    fila[COLUMNA_TENENCIA] = "Sí"

    vista = await cargar_planilla(file=_subir([fila]), dry_run=True, pool=pool, user=usuario)
    assert vista["cambian"] == 1 and vista["aplicado"] is False
    assert await conexion_revertida.fetchval(
        "SELECT status FROM public.compliance_records WHERE id = $1::uuid",
        objetivo["id"]) == "MISSING", "el dry_run escribió"


async def test_tambien_acepta_csv(conexion_revertida):
    """Baja XLSX porque es lo cómodo, pero alguien la va a pasar por Google
    Sheets y a devolverla como CSV. Rechazarla ahí es hacerle perder el
    trabajo por el formato."""
    import csv as _csv
    pool = PoolDeUnaConexion(conexion_revertida)
    usuario = await _usuario_real(conexion_revertida)
    empresa = await _empresa_activa(conexion_revertida)
    objetivo = (await _registros_de(conexion_revertida, empresa, con_vencimiento=False))[0]

    fila = await _fila(pool, objetivo["id"])
    fila[COLUMNA_TENENCIA] = "si"          # sin tilde, como lo escribe cualquiera
    buffer = io.StringIO()
    escritor = _csv.writer(buffer, delimiter=";", lineterminator="\r\n")
    claves = [c["csv_key"] for c in COLUMNAS]
    escritor.writerow(claves)
    escritor.writerow([fila[k] for k in claves])
    archivo = UploadFile(file=io.BytesIO(buffer.getvalue().encode("utf-8-sig")),
                         filename="certificacion.csv")

    resultado = await cargar_planilla(file=archivo, dry_run=False, pool=pool, user=usuario)
    assert resultado["cambian"] == 1, resultado["errores"]


async def test_una_tenencia_que_no_se_entiende_se_avisa(conexion_revertida):
    """Con None querría decir "vacío" y la fila se ignoraría en silencio."""
    pool = PoolDeUnaConexion(conexion_revertida)
    usuario = await _usuario_real(conexion_revertida)
    empresa = await _empresa_activa(conexion_revertida)
    objetivo = (await _registros_de(conexion_revertida, empresa, con_vencimiento=False))[0]

    fila = await _fila(pool, objetivo["id"])
    fila[COLUMNA_TENENCIA] = "mas o menos"

    vista = await cargar_planilla(file=_subir([fila]), dry_run=True, pool=pool, user=usuario)
    assert vista["total_errores"] == 1 and vista["vacias"] == 0
    assert "no se entiende" in vista["errores"][0]["error"].lower()


async def test_se_escribe_en_una_sola_pasada_y_queda_auditado(conexion_revertida):
    """`trg_refresh_view_on_compliance` refresca una vista materializada POR
    STATEMENT sobre compliance_records: de a una, miles de filas son miles de
    refrescos. Y la auditoría es por CAMPO: una fila que mueve los dos ejes son
    dos hechos distintos."""
    pool = PoolDeUnaConexion(conexion_revertida)
    usuario = await _usuario_real(conexion_revertida)
    empresa = await _empresa_activa(conexion_revertida)
    tres = (await _registros_de(conexion_revertida, empresa, con_vencimiento=True))[:3]
    assert len(tres) == 3

    filas = []
    for numero, registro in enumerate(tres, start=1):
        fila = await _fila(pool, registro["id"])
        fila[COLUMNA_TENENCIA] = "Sí"      # mueve evidencia
        fila[COLUMNA_VENCIMIENTO] = (      # y vigencia
            date.today() + timedelta(days=30 * numero)).strftime("%d-%m-%Y")
        filas.append(fila)

    antes = await conexion_revertida.fetchval(
        "SELECT count(*) FROM public.audit_log WHERE source = $1", FUENTE_AUDITORIA)
    resultado = await cargar_planilla(file=_subir(filas), dry_run=False, pool=pool, user=usuario)
    assert resultado["cambian"] == 3 and resultado["recibidos"] == 3 and resultado["fechas"] == 3

    despues = await conexion_revertida.fetchval(
        "SELECT count(*) FROM public.audit_log WHERE source = $1", FUENTE_AUDITORIA)
    assert despues - antes == 6, "3 filas × 2 ejes = 6 hechos auditados"


async def test_el_resumen_separa_los_dos_ejes(conexion_revertida):
    pool = PoolDeUnaConexion(conexion_revertida)
    await _empresa_activa(conexion_revertida)
    resumen = await resumen_de_planilla(alcance="activas", pool=pool, _=None)
    assert resumen["filas"] == resumen["con_vencimiento"] + resumen["solo_tenencia"]
    assert resumen["solo_tenencia"] > 0, (
        "si esto da cero, la planilla volvió a traer sólo los que vencen"
    )
    assert set(resumen["por_entidad"]) <= {"Empresa", "Conductor", "Vehículo"}


# ── Forma ────────────────────────────────────────────────────────────────────


@pytest.mark.parametrize("sql, argumentos_que_pasa_el_endpoint",
                         [(SQL_APLICAR, 4), (SQL_AUDITAR, 7)])
def test_los_placeholders_coinciden_con_los_argumentos(sql, argumentos_que_pasa_el_endpoint):
    """Sustituir $n por literales para probar una consulta no prueba el binding.
    Un placeholder de más o de menos revienta recién en producción, con un
    'the server expects N arguments' que no dice cuál falta."""
    assert max(int(n) for n in re.findall(r"\$(\d+)", sql)) == argumentos_que_pasa_el_endpoint


def test_la_llave_baja_bloqueada():
    """El id_registro es lo que devuelve cada fila a su lugar. Una llave
    editable por accidente es una fila aplicada al documento equivocado."""
    from openpyxl import load_workbook
    hoja = load_workbook(io.BytesIO(_planilla_a_xlsx(
        [{c["csv_key"]: "x" for c in COLUMNAS}]))).active
    assert hoja.protection.sheet is True
    for columna, definicion in enumerate(COLUMNAS, start=1):
        celda = hoja.cell(row=2, column=columna)
        assert celda.protection.locked is not definicion["editable"], definicion["csv_key"]


async def test_un_ida_y_vuelta_sin_tocar_nada_no_cambia_nada(conexion_revertida):
    """La planilla baja con las dos columnas PRE-LLENADAS. Si un estado que la
    columna de tenencia no puede expresar —vencido, rechazado— se mapeara a sí o
    a no, bajarla y volver a subirla SIN TOCARLA los convertiría en silencio: un
    rechazado pasaría a faltante, un vencido a recibido. Hoy son 0 filas en la
    base; el test existe para que sigan siendo 0 problemas."""
    pool = PoolDeUnaConexion(conexion_revertida)
    usuario = await _usuario_real(conexion_revertida)
    empresa = await _empresa_activa(conexion_revertida)
    registros = await _registros_de(conexion_revertida, empresa, con_vencimiento=True)

    # Se fabrican los estados que hoy no existen en la base. Los dos llevan
    # fecha pasada a propósito: el predicado compartido de "pendiente" NO
    # incluye REJECTED, así que un rechazado sin fecha no aparece en ninguna
    # cola del módulo — es preexistente y no se toca desde acá.
    vencio = date.today() - timedelta(days=10)
    for registro, estado in ((registros[0], "REJECTED"), (registros[1], "EXPIRED")):
        await conexion_revertida.execute(
            "UPDATE public.compliance_records SET status = $2, expiration_date = $3 "
            "WHERE id = $1::uuid",
            registro["id"], estado, vencio)

    ids = [r["id"] for r in registros[:2]]
    filas = [await _fila(pool, i) for i in ids]          # tal cual bajan
    resultado = await cargar_planilla(file=_subir(filas), dry_run=True, pool=pool, user=usuario)

    assert resultado["cambian"] == 0, (
        "bajar la planilla y volver a subirla sin tocarla cambió algo: "
        f"{resultado}"
    )
    # Y la razón por la que no cambió: esos dos estados bajan con la celda de
    # tenencia EN BLANCO, porque un sí/no no puede representarlos.
    for fila in filas:
        assert fila[COLUMNA_TENENCIA] == "", (
            f'{fila["estado_actual"]} bajó como "{fila[COLUMNA_TENENCIA]}"'
        )
