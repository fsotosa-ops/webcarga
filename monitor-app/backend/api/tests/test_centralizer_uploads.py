"""Tests de `routers/centralizer_uploads.py` (Task 3). Dos niveles:

1. Tests HTTP end-to-end (`make_client` + `TestClient`, mismo patrón que
   `tests/test_transporters_relational.py`) para el ciclo
   POST/approve/reject/apply y sus guardas de estado (409/403/404).
2. Tests unitarios directos sobre los helpers de apply (`_apply_transporter`,
   `_apply_driver_or_vehicle`, `_apply_field_diffs`) con una conexión fake
   liviana (`FakeConn`) en vez de encadenar `side_effect` exactos sobre un
   AsyncMock genérico — el orden/cantidad de llamadas fetch/fetchval/fetchrow
   dentro de una transacción completa de apply (compute_diff + inserts +
   upserts de documentos) es demasiado largo para mantener a mano sin que
   quede frágil ante cualquier reordenamiento interno no relevante para el
   comportamiento bajo prueba.
"""
import json
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

from app.auth import get_current_user, get_supabase, require_admin, require_editor
from app.db import get_pool
from app.routers.centralizer_uploads import (
    _apply_diff,
    _apply_driver_or_vehicle,
    _apply_field_diffs,
    _apply_transporter,
    router,
)

USER_ID = "11111111-1111-1111-1111-111111111111"
UPLOAD_ID = "aaaaaaaa-0000-0000-0000-000000000099"
TID = "aaaaaaaa-0000-0000-0000-000000000001"
NEW_TID = "aaaaaaaa-0000-0000-0000-000000000002"
OLD_TID = "aaaaaaaa-0000-0000-0000-000000000003"
DID = "bbbbbbbb-0000-0000-0000-000000000002"
USER = {"sub": USER_ID, "email": "admin@webcarga.cl", "role": "admin"}


def make_client(pool, role="admin", enforce_roles=False, supabase=None):
    app = FastAPI()
    app.include_router(router, prefix="/api/v1")
    app.dependency_overrides[get_pool] = lambda: pool
    app.dependency_overrides[get_supabase] = lambda: supabase or MagicMock()
    user = {"sub": USER_ID, "email": "operador@webcarga.cl", "role": role}
    app.dependency_overrides[get_current_user] = lambda: user
    if not enforce_roles:
        app.dependency_overrides[require_editor] = lambda: user
        app.dependency_overrides[require_admin] = lambda: user
    return TestClient(app)


def _fixture_bytes() -> bytes:
    with open("tests/fixtures/centralizer_sample.xlsx", "rb") as f:
        return f.read()


# ── _download_and_parse: helper compartido por apply y GET /{id} ─────────

def test_download_and_parse_returns_parsed_workbook():
    from app.routers.centralizer_uploads import _download_and_parse
    supabase = MagicMock()
    supabase.storage.from_.return_value.download.return_value = _fixture_bytes()

    parsed = _download_and_parse(supabase, "centralizer-uploads/x.xlsx")

    assert parsed["sheet_summary"] == {"Empresas": 2, "Conductores": 3, "Vehiculos_Equipos": 3}
    supabase.storage.from_.assert_called_with("compliance-docs")


def test_download_and_parse_storage_error_raises_502():
    from fastapi import HTTPException
    from app.routers.centralizer_uploads import _download_and_parse
    supabase = MagicMock()
    supabase.storage.from_.return_value.download.side_effect = Exception("boom")

    with pytest.raises(HTTPException) as exc:
        _download_and_parse(supabase, "x.xlsx")
    assert exc.value.status_code == 502


def test_download_and_parse_missing_sheet_raises_422():
    from io import BytesIO
    from openpyxl import Workbook
    from fastapi import HTTPException
    from app.routers.centralizer_uploads import _download_and_parse

    wb = Workbook()
    wb.active.title = "Empresas"  # falta Conductores/Vehiculos_Equipos
    buf = BytesIO()
    wb.save(buf)

    supabase = MagicMock()
    supabase.storage.from_.return_value.download.return_value = buf.getvalue()

    with pytest.raises(HTTPException) as exc:
        _download_and_parse(supabase, "x.xlsx")
    assert exc.value.status_code == 422


# ── GET /{id} y GET / — diff recalculado + nombres de profiles ───────────

def _upload_row(**overrides):
    row = {
        "id": UPLOAD_ID, "upload_kind": "centralizer", "file_name": "centralizador.xlsx",
        "storage_path": "centralizer-uploads/x.xlsx", "uploaded_by": USER_ID,
        "uploaded_at": "2026-07-13T00:00:00", "status": "previewed",
        "sheet_summary": {"Empresas": 2, "Conductores": 3, "Vehiculos_Equipos": 3},
        "parse_errors": [], "approved_by": None, "approved_at": None, "applied_at": None,
        "rejected_by": None, "rejected_at": None, "rejection_reason": None,
        "created_at": "2026-07-13T00:00:00",
        "uploaded_by_name": "Ana Pérez", "approved_by_name": None, "rejected_by_name": None,
    }
    row.update(overrides)
    return row


def test_get_upload_previewed_includes_recomputed_diff():
    pool = AsyncMock()
    pool.fetchrow.return_value = _upload_row(status="previewed")
    pool.fetch.side_effect = [[], [], [], []]  # _load_extra_mappings, luego compute_diff: sin matches existentes

    supabase = MagicMock()
    supabase.storage.from_.return_value.download.return_value = _fixture_bytes()

    client = make_client(pool, supabase=supabase)
    res = client.get(f"/api/v1/centralizer-uploads/{UPLOAD_ID}")

    assert res.status_code == 200, res.text
    data = res.json()["data"]
    assert data["uploaded_by_name"] == "Ana Pérez"
    assert len(data["diff"]["transporters"]) == 2
    supabase.storage.from_.assert_called_with("compliance-docs")


def test_get_upload_failed_status_has_no_diff_and_skips_storage():
    pool = AsyncMock()
    pool.fetchrow.return_value = _upload_row(status="failed", storage_path=None)
    supabase = MagicMock()

    client = make_client(pool, supabase=supabase)
    res = client.get(f"/api/v1/centralizer-uploads/{UPLOAD_ID}")

    assert res.status_code == 200
    assert res.json()["data"]["diff"] is None
    supabase.storage.from_.assert_not_called()


def test_get_upload_pending_mapping_returns_unresolved_columns_not_diff():
    from io import BytesIO
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = 'Empresas'
    ws.append(['Nombre / Razón Social', 'RUT', 'DV', 'Otra Columna Nueva'])
    ws.append(['Test SPA', '99999013', '1', 'x'])
    wb.create_sheet('Conductores')
    wb.create_sheet('Vehiculos_Equipos')
    buf = BytesIO()
    wb.save(buf)

    pool = AsyncMock()
    pool.fetchrow.return_value = _upload_row(status="pending_mapping", storage_path="centralizer-uploads/x.xlsx")
    pool.fetch.return_value = []  # _load_extra_mappings: sin mapeos guardados

    supabase = MagicMock()
    supabase.storage.from_.return_value.download.return_value = buf.getvalue()

    client = make_client(pool, supabase=supabase)
    res = client.get(f"/api/v1/centralizer-uploads/{UPLOAD_ID}")

    assert res.status_code == 200, res.text
    data = res.json()["data"]
    assert data["diff"] is None
    assert data["unresolved_columns"] == [{"sheet": "Empresas", "header": "Otra Columna Nueva"}]


def test_get_upload_not_found_returns_404():
    pool = AsyncMock()
    pool.fetchrow.return_value = None
    client = make_client(pool)
    res = client.get(f"/api/v1/centralizer-uploads/{UPLOAD_ID}")
    assert res.status_code == 404


def test_list_uploads_includes_display_names():
    pool = AsyncMock()
    pool.fetch.return_value = [_upload_row()]
    pool.fetchval.return_value = 1
    client = make_client(pool)
    res = client.get("/api/v1/centralizer-uploads")
    assert res.status_code == 200
    assert res.json()["data"][0]["uploaded_by_name"] == "Ana Pérez"


# ── FakeConn: conexión de transacción con estado mínimo para apply ────────

class _NullCtx:
    async def __aenter__(self):
        return None

    async def __aexit__(self, *exc):
        return False


class FakeConn:
    """Simula asyncpg.Connection dentro de `async with pool.acquire() as conn:
    async with conn.transaction():` para un escenario de apply "todo nuevo"
    (base de datos vacía): toda query de existencia (`app.transporters`/
    `drivers`/`vehicles`/`*_documents`) devuelve vacío, todo INSERT...RETURNING
    id devuelve un uuid incremental, y `compliance_doc_catalog` siempre valida
    el doc_code pedido."""

    def __init__(self, upload_status="approved"):
        self.upload_status = upload_status
        self.executed: list[tuple] = []
        self._next = 0

    def _new_id(self) -> str:
        self._next += 1
        return f"cccccccc-0000-0000-0000-{self._next:012d}"

    def transaction(self):
        return _NullCtx()

    async def execute(self, sql, *args):
        self.executed.append((sql, args))
        return "UPDATE 1"

    async def fetch(self, sql, *args):
        return []  # compute_diff: sin matches existentes, sin documentos existentes

    async def fetchval(self, sql, *args):
        if "SELECT status FROM app.centralizer_uploads" in sql:
            return self.upload_status
        if "SELECT doc_code FROM app.compliance_doc_catalog" in sql:
            return args[0]
        if sql.strip().startswith("INSERT INTO app.transporters"):
            return self._new_id()
        if sql.strip().startswith("INSERT INTO app.drivers"):
            return self._new_id()
        if sql.strip().startswith("INSERT INTO app.vehicles"):
            return self._new_id()
        raise AssertionError(f"fetchval inesperado en FakeConn: {sql}")

    async def fetchrow(self, sql, *args):
        if "WITH old AS" in sql:
            return None
        return {}  # RETURNING * de _upsert_document — valor de retorno no usado por el caller


def _acquire_ctx(conn):
    ctx = MagicMock()
    ctx.__aenter__ = AsyncMock(return_value=conn)
    ctx.__aexit__ = AsyncMock(return_value=False)
    return ctx


# ── POST /centralizer-uploads: sube + parsea + diff ────────────────────────

def test_upload_and_preview_parses_and_returns_structured_diff():
    pool = AsyncMock()
    pool.fetch.side_effect = [[], [], [], []]  # _load_extra_mappings, luego transporters/drivers/vehicles existentes: ninguno
    pool.fetchval.return_value = UPLOAD_ID

    supabase = MagicMock()
    client = make_client(pool, supabase=supabase)

    res = client.post(
        "/api/v1/centralizer-uploads",
        files={"file": (
            "centralizer_sample.xlsx", _fixture_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )},
    )

    assert res.status_code == 200, res.text
    data = res.json()
    assert data["upload_id"] == UPLOAD_ID
    assert data["sheet_summary"] == {"Empresas": 2, "Conductores": 3, "Vehiculos_Equipos": 3}
    assert data["parse_errors"] == []
    diff = data["diff"]
    assert len(diff["transporters"]) == 2
    assert len(diff["drivers"]) == 3
    assert len(diff["vehicles"]) == 3
    assert all(d["change_type"] == "new" for d in diff["transporters"])

    supabase.storage.from_.assert_called_with("compliance-docs")
    insert_sql = pool.fetchval.call_args.args[0]
    assert "INSERT INTO app.centralizer_uploads" in insert_sql
    assert "'previewed'" in insert_sql


def test_upload_missing_required_sheet_persists_failed_row():
    # Falta la hoja "Conductores"/"Vehiculos_Equipos" -> sigue siendo un
    # fallo duro (no es un problema de mapeo de columnas resoluble por un
    # admin, es una estructura de archivo inválida).
    from io import BytesIO
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas"
    ws.append(["Nombre / Razón Social", "RUT", "DV"])
    ws.append(["Test SPA", "11111111", "1"])
    buf = BytesIO()
    wb.save(buf)

    pool = AsyncMock()
    pool.fetch.return_value = []  # _load_extra_mappings
    pool.fetchval.return_value = "dddddddd-0000-0000-0000-000000000001"
    client = make_client(pool)

    res = client.post(
        "/api/v1/centralizer-uploads",
        files={"file": (
            "bad.xlsx", buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )},
    )

    assert res.status_code == 422
    assert "upload_id" in res.json()["detail"]
    insert_sql = pool.fetchval.call_args.args[0]
    assert "'failed'" in insert_sql


def test_upload_with_unmapped_column_returns_pending_mapping_not_422():
    # Comportamiento nuevo: una columna sin mapear ya NO bloquea el upload
    # completo con un 422 — cae en 'pending_mapping' para que un admin la
    # resuelva (mapear/crear/ignorar), ver routers/centralizer_uploads.py.
    from io import BytesIO
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas"
    ws.append(["Nombre / Razón Social", "RUT", "DV", "Columna Sin Mapeo"])
    ws.append(["Test SPA", "11111111", "1", "x"])
    wb.create_sheet("Conductores")
    wb.create_sheet("Vehiculos_Equipos")
    buf = BytesIO()
    wb.save(buf)

    pool = AsyncMock()
    pool.fetch.return_value = []  # _load_extra_mappings: sin mapeos guardados
    pool.fetchval.return_value = "dddddddd-0000-0000-0000-000000000001"
    client = make_client(pool)

    res = client.post(
        "/api/v1/centralizer-uploads",
        files={"file": (
            "bad.xlsx", buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )},
    )

    assert res.status_code == 200, res.text
    data = res.json()
    assert data["status"] == "pending_mapping"
    assert data["unresolved_columns"] == [{"sheet": "Empresas", "header": "Columna Sin Mapeo"}]
    insert_sql = pool.fetchval.call_args.args[0]
    assert "'pending_mapping'" in insert_sql


def test_upload_with_saved_mapping_present_still_parses_normally():
    # Confirma que el flujo normal (sin columnas nuevas del todo) sigue
    # funcionando cuando YA hay mapeos guardados en la tabla (irrelevantes
    # para este archivo en particular) — no deben interferir.
    pool = AsyncMock()
    saved_mapping = [{"sheet_name": "Empresas", "excel_header": "Cuenta Banco Empresa", "doc_code": "cuenta_banco_empresa"}]
    pool.fetch.side_effect = [saved_mapping, [], [], []]  # _load_extra_mappings, luego compute_diff x3
    pool.fetchval.return_value = "eeeeeeee-0000-0000-0000-000000000003"

    client = make_client(pool)
    res = client.post(
        "/api/v1/centralizer-uploads",
        files={"file": (
            "centralizer_sample.xlsx", _fixture_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )},
    )
    assert res.status_code == 200, res.text
    assert res.json().get("status") != "pending_mapping"
    assert "diff" in res.json()


# ── approve ──────────────────────────────────────────────────────────────

def test_approve_requires_admin_403_with_editor_role():
    pool = AsyncMock()
    client = make_client(pool, role="editor", enforce_roles=True)
    res = client.post(f"/api/v1/centralizer-uploads/{UPLOAD_ID}/approve")
    assert res.status_code == 403


def test_approve_success_from_previewed():
    pool = AsyncMock()
    pool.execute.return_value = "UPDATE 1"
    client = make_client(pool)
    res = client.post(f"/api/v1/centralizer-uploads/{UPLOAD_ID}/approve")
    assert res.status_code == 200
    assert res.json()["status"] == "approved"


def test_approve_wrong_status_returns_409():
    pool = AsyncMock()
    pool.execute.return_value = "UPDATE 0"
    pool.fetchval.return_value = "applied"
    client = make_client(pool)
    res = client.post(f"/api/v1/centralizer-uploads/{UPLOAD_ID}/approve")
    assert res.status_code == 409


def test_approve_not_found_returns_404():
    pool = AsyncMock()
    pool.execute.return_value = "UPDATE 0"
    pool.fetchval.return_value = None
    client = make_client(pool)
    res = client.post(f"/api/v1/centralizer-uploads/{UPLOAD_ID}/approve")
    assert res.status_code == 404


# ── reject ───────────────────────────────────────────────────────────────

def test_reject_success():
    pool = AsyncMock()
    pool.execute.return_value = "UPDATE 1"
    client = make_client(pool)
    res = client.post(f"/api/v1/centralizer-uploads/{UPLOAD_ID}/reject", json={"reason": "datos incorrectos"})
    assert res.status_code == 200
    assert res.json()["status"] == "rejected"


# ── apply: guardas de estado + idempotencia ─────────────────────────────

def test_apply_fails_409_if_not_approved():
    pool = AsyncMock()
    pool.fetchrow.return_value = {"id": UPLOAD_ID, "status": "previewed", "storage_path": "x"}
    client = make_client(pool)
    res = client.post(f"/api/v1/centralizer-uploads/{UPLOAD_ID}/apply")
    assert res.status_code == 409
    pool.acquire.assert_not_called()


def test_apply_on_already_applied_upload_is_rejected_not_reapplied():
    """Idempotencia: un segundo apply sobre un upload que ya está 'applied'
    debe rechazar con 409, no re-aplicar en silencio (el status nunca vuelve
    a 'approved' después de 'applied')."""
    pool = AsyncMock()
    pool.fetchrow.return_value = {"id": UPLOAD_ID, "status": "applied", "storage_path": "x"}
    client = make_client(pool)
    res = client.post(f"/api/v1/centralizer-uploads/{UPLOAD_ID}/apply")
    assert res.status_code == 409
    pool.acquire.assert_not_called()


def test_apply_not_found_returns_404():
    pool = AsyncMock()
    pool.fetchrow.return_value = None
    client = make_client(pool)
    res = client.post(f"/api/v1/centralizer-uploads/{UPLOAD_ID}/apply")
    assert res.status_code == 404


# ── apply: flujo completo exitoso (base vacía -> todo 'new') ───────────────

def test_apply_success_end_to_end_all_new_marks_matched_and_applied():
    pool = AsyncMock()
    pool.fetchrow.return_value = {"id": UPLOAD_ID, "status": "approved", "storage_path": "centralizer-uploads/x.xlsx"}
    pool.fetch.return_value = []  # _load_extra_mappings: sin mapeos guardados
    conn = FakeConn(upload_status="approved")
    pool.acquire = MagicMock(return_value=_acquire_ctx(conn))

    supabase = MagicMock()
    supabase.storage.from_.return_value.download.return_value = _fixture_bytes()

    client = make_client(pool, supabase=supabase)
    res = client.post(f"/api/v1/centralizer-uploads/{UPLOAD_ID}/apply")

    assert res.status_code == 200, res.text
    data = res.json()
    assert data["status"] == "applied"
    assert data["matched_transporters"] == 2  # las 2 empresas del fixture

    applied_sql = next(sql for sql, _ in conn.executed if "SET status = 'applied'" in sql)
    assert applied_sql

    matched_call = next((sql, args) for sql, args in conn.executed if "last_matched_upload_id" in sql)
    assert matched_call[1][0] == UPLOAD_ID
    assert len(matched_call[1][1]) == 2  # 2 transporters matcheados


def test_apply_race_second_concurrent_apply_rejected_inside_transaction():
    """Si dos requests pasan el chequeo inicial (status=='approved') antes de
    que cualquiera commitee, el re-chequeo DENTRO de la transacción (tras el
    advisory lock) debe rechazar al que llega después."""
    pool = AsyncMock()
    pool.fetchrow.return_value = {"id": UPLOAD_ID, "status": "approved", "storage_path": "x.xlsx"}
    pool.fetch.return_value = []  # _load_extra_mappings: sin mapeos guardados
    conn = FakeConn(upload_status="applied")  # otro apply ya commiteó mientras esperábamos el lock
    pool.acquire = MagicMock(return_value=_acquire_ctx(conn))

    supabase = MagicMock()
    supabase.storage.from_.return_value.download.return_value = _fixture_bytes()

    client = make_client(pool, supabase=supabase)
    res = client.post(f"/api/v1/centralizer-uploads/{UPLOAD_ID}/apply")

    assert res.status_code == 409
    assert not any("SET status = 'applied'" in sql for sql, _ in conn.executed)


# ── Unit: helpers de apply respetan conflict / manual_override ────────────

@pytest.mark.asyncio
async def test_apply_transporter_conflict_is_matched_but_no_write():
    conn = FakeConn()
    ed = {
        "entity_key": "99999099", "match_method": "rut", "existing_id": TID,
        "change_type": "conflict",
        "field_diffs": [{"field": "business_name", "old": "A", "new": "B", "conflict": True}],
        "conflict_reason": "manually_edited_field",
        "parsed_row": {"rut": "99999099", "business_name": "B", "documents": {}},
    }
    result = await _apply_transporter(conn, ed, USER)
    assert result == TID
    assert conn.executed == []


@pytest.mark.asyncio
async def test_apply_driver_or_vehicle_conflict_skips_entirely():
    conn = FakeConn()
    ed = {
        "entity_key": "22222222", "match_method": "rut", "existing_id": DID,
        "change_type": "conflict",
        "field_diffs": [{"field": "full_name", "old": "A", "new": "B", "conflict": True}],
        "conflict_reason": "manually_edited_field",
        "parsed_row": {"rut": "22222222", "transporter_rut": "99999099", "documents": {}},
    }
    await _apply_driver_or_vehicle(conn, ed, "driver", USER, {"99999099": TID})
    assert conn.executed == []


@pytest.mark.asyncio
async def test_apply_field_diffs_skips_conflicts_and_routes_docs_with_manual_override_false():
    conn = AsyncMock()
    field_diffs = [
        {"field": "business_name", "old": "A", "new": "B", "conflict": False},
        {"field": "documents.rol_sii", "old": "pendiente", "new": "ok", "conflict": False},
        {"field": "documents.f43", "old": None, "new": "ok", "conflict": True},
    ]
    with patch(
        "app.routers.centralizer_uploads._upsert_document", new=AsyncMock(),
    ) as mock_upsert:
        await _apply_field_diffs(conn, "transporter", TID, field_diffs, USER)

    mock_upsert.assert_called_once_with(
        conn, "transporter", TID, "rol_sii", {"status": "ok", "manual_override": False}, USER_ID,
    )
    update_call = conn.execute.call_args
    assert "business_name = $2" in update_call.args[0]
    assert update_call.args[1] == TID
    assert update_call.args[2] == "B"


@pytest.mark.asyncio
async def test_driver_reassignment_logs_transfer_audit_only_when_old_transporter_existed():
    conn = AsyncMock()
    conn.fetchrow.return_value = {"old_transporter_id": OLD_TID}
    ed = {
        "entity_key": "22222222", "match_method": "rut", "existing_id": DID,
        "change_type": "unchanged", "field_diffs": [], "conflict_reason": None,
        "parsed_row": {"rut": "22222222", "transporter_rut": "99999099", "documents": {}},
    }
    await _apply_driver_or_vehicle(conn, ed, "driver", USER, {"99999099": NEW_TID})

    audit_call = next(c for c in conn.execute.call_args_list if "audit_log" in c.args[0])
    assert audit_call.args[1] == USER_ID
    assert audit_call.args[2] == "driver"
    assert audit_call.args[3] == DID
    assert json.loads(audit_call.args[4]) == {"transporter_id": OLD_TID}
    assert json.loads(audit_call.args[5]) == {"transporter_id": NEW_TID}


@pytest.mark.asyncio
async def test_driver_new_assignment_does_not_log_transfer_audit():
    conn = AsyncMock()
    conn.fetchrow.return_value = {"old_transporter_id": None}
    ed = {
        "entity_key": "22222222", "match_method": "rut", "existing_id": DID,
        "change_type": "unchanged", "field_diffs": [], "conflict_reason": None,
        "parsed_row": {"rut": "22222222", "transporter_rut": "99999099", "documents": {}},
    }
    await _apply_driver_or_vehicle(conn, ed, "driver", USER, {"99999099": NEW_TID})

    assert not any("audit_log" in c.args[0] for c in conn.execute.call_args_list)


@pytest.mark.asyncio
async def test_apply_diff_new_transporter_and_driver_updates_last_matched():
    conn = FakeConn()
    diff = {
        "transporters": [{
            "entity_key": "99999001", "match_method": None, "existing_id": None,
            "change_type": "new", "field_diffs": [
                {"field": "business_name", "old": None, "new": "Transportes Prueba SPA", "conflict": False},
            ],
            "conflict_reason": None,
            "parsed_row": {"rut": "99999001", "dv": "1", "rut_dv_valid": True,
                           "business_name": "Transportes Prueba SPA", "documents": {}},
        }],
        "drivers": [{
            "entity_key": "11111111", "match_method": None, "existing_id": None,
            "change_type": "new", "field_diffs": [
                {"field": "full_name", "old": None, "new": "Juan Test", "conflict": False},
            ],
            "conflict_reason": None,
            "parsed_row": {"rut": "11111111", "dv": "1", "rut_dv_valid": True,
                           "transporter_rut": "99999001", "full_name": "Juan Test", "documents": {}},
        }],
        "vehicles": [],
        "parse_errors": [],
    }
    matched = await _apply_diff(conn, diff, UPLOAD_ID, USER)
    assert len(matched) == 1

    matched_call = next((sql, args) for sql, args in conn.executed if "last_matched_upload_id" in sql)
    assert matched_call[1][0] == UPLOAD_ID
    assert list(matched_call[1][1]) == list(matched)
