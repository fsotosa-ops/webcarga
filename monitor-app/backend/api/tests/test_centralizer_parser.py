"""Tests para app/services/centralizer_parser.py.

El fixture `tests/fixtures/centralizer_sample.xlsx` es 100% sintético:
mismos headers exactos que el Excel real de producción (importados
directamente desde los dicts *_COLUMNS del módulo al generarlo, ver
script de generación descartable usado una vez), pero con RUTs y nombres
de empresa/conductor inventados. Los RUTs usados (99999001, 99999002,
99999101, 99999102, 99999103) fueron verificados contra
app.transporters / app.drivers en Supabase (0 filas) antes de fijarlos,
usando un patrón fuera de rango (prefijo 99999) para minimizar el riesgo
de colisión futura con RUTs reales.
"""
import pytest
from datetime import date
from app.services.centralizer_parser import (
    normalize_rut,
    rut_dv,
    parse_centralizer_date,
    map_doc_status,
    parse_centralizer_workbook,
)


def test_normalize_rut_strips_dots_and_dv():
    assert normalize_rut('12.345.678-9') == '12345678'
    assert normalize_rut('12345678-9') == '12345678'
    assert normalize_rut('12345678') == '12345678'


def test_rut_dv_matches_known_valid_ruts():
    # RUTs reales válidos conocidos en Chile (dígito verificador módulo 11) —
    # institucionales / de dominio público, no personales ni de una empresa
    # de transporte real. Mismos 3 RUTs ya verificados a mano en el
    # docstring de centralizer_parser.py (SII, BancoEstado, Universidad de
    # Chile). Nota: el ejemplo del brief original (76.086.428-3) tenía un DV
    # incorrecto — el cálculo módulo-11 a mano da '5', no '3', para ese RUT;
    # se reemplazó por los 3 ejemplos ya verificados en el módulo.
    assert rut_dv('60803000') == 'K'
    assert rut_dv('97030000') == '7'
    assert rut_dv('60910000') == '1'


def test_parse_centralizer_date_valid_and_invalid():
    assert parse_centralizer_date('15/03/2027') == date(2027, 3, 15)
    assert parse_centralizer_date('') is None
    assert parse_centralizer_date('no-es-fecha') is None
    assert parse_centralizer_date(None) is None


def test_map_doc_status_known_values():
    assert map_doc_status('OK') == 'ok'
    assert map_doc_status('Pendiente') == 'pendiente'
    assert map_doc_status('Factible') == 'factible'
    assert map_doc_status('N/A') == 'n_a'
    assert map_doc_status('') is None
    assert map_doc_status('valor-desconocido') is None


def test_parse_centralizer_workbook_reads_all_3_sheets():
    with open('tests/fixtures/centralizer_sample.xlsx', 'rb') as f:
        result = parse_centralizer_workbook(f.read())

    # 3 filas en Empresas colapsan a 2 transporters (dedupe multi-cliente)
    assert result['sheet_summary']['Empresas'] == 2
    assert result['sheet_summary']['Conductores'] == 3
    assert result['sheet_summary']['Vehiculos_Equipos'] == 3
    assert len(result['transporters']) == result['sheet_summary']['Empresas']
    assert len(result['drivers']) == result['sheet_summary']['Conductores']
    assert len(result['vehicles']) == result['sheet_summary']['Vehiculos_Equipos']
    assert result['parse_errors'] == []

    t = result['transporters'][0]
    assert 'rut' in t and 'dv' in t and 'business_name' in t
    assert 'documents' in t and isinstance(t['documents'], dict)
    # doc_code mapeado, no el nombre de columna en español crudo:
    assert 'rol_sii' in t['documents']

    d = result['drivers'][0]
    assert 'transporter_rut' in d and 'full_name' in d and 'rut' in d
    assert 'documents' in d and 'copia_ci' in d['documents']

    v = result['vehicles'][0]
    assert 'transporter_rut' in v and 'plate' in v and 'kind' in v
    # los vehículos no traen rut/dv propios (se identifican por patente)
    assert 'rut' not in v and 'dv' not in v


def test_parse_centralizer_workbook_maps_anexo_repleg_gc_column():
    # Columna real de producción no mapeada hasta ahora ("ANEXO RepLeg (GC)")
    # — anexo firmado por el representante legal, pedido por generadores de
    # carga como Walmart. Confirma que mapea a documents['anexo_repleg_gc']
    # en vez de levantar "columna no mapeada".
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = 'Empresas'
    ws.append(['Nombre / Razón Social', 'RUT', 'DV', 'ANEXO RepLeg (GC)'])
    ws.append(['Transportes Prueba SPA', '99999007', '5', 'OK'])
    wb.create_sheet('Conductores')
    wb.create_sheet('Vehiculos_Equipos')
    buf = BytesIO()
    wb.save(buf)

    result = parse_centralizer_workbook(buf.getvalue())

    assert result['parse_errors'] == []
    assert result['transporters'][0]['documents']['anexo_repleg_gc'] == 'ok'


def test_parse_centralizer_workbook_dedupes_multi_cliente_transporter():
    with open('tests/fixtures/centralizer_sample.xlsx', 'rb') as f:
        result = parse_centralizer_workbook(f.read())

    by_rut = {t['rut']: t for t in result['transporters']}
    assert set(by_rut.keys()) == {'99999001', '99999002'}

    dedup = by_rut['99999001']
    assert dedup['business_name'] == 'Transportes Prueba Uno SPA'
    assert len(dedup['clients']) == 2
    assert {c['avance_80_20'] for c in dedup['clients']} == {80, 60}

    single = by_rut['99999002']
    assert len(single['clients']) == 1


def test_parse_centralizer_workbook_derives_status_from_expiry_date():
    with open('tests/fixtures/centralizer_sample.xlsx', 'rb') as f:
        result = parse_centralizer_workbook(f.read())

    driver = next(d for d in result['drivers'] if d['rut'] == '99999101')
    # Licencia (Vencimiento) = 01/01/2020, ya vencida -> 'actualizar'
    assert driver['license_expiry'] == date(2020, 1, 1)
    assert driver['documents']['licencia'] == 'actualizar'
    # Copia C.I (Vencimiento) = 15/03/2027, futura -> 'ok'
    assert driver['documents']['copia_ci'] == 'ok'


def test_parse_centralizer_workbook_rejects_unmapped_column():
    # Construir un xlsx mínimo en memoria con una columna que no existe en el
    # mapeo (usar openpyxl.Workbook() directo, no el fixture) y confirmar que
    # parse_centralizer_workbook levanta una excepción clara en vez de
    # ignorar la columna en silencio.
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = 'Empresas'
    ws.append(['Nombre / Razón Social', 'RUT', 'DV', 'Columna Inventada Sin Mapeo'])
    ws.append(['Test SPA', '11111111', '1', 'x'])
    wb.create_sheet('Conductores')
    wb.create_sheet('Vehiculos_Equipos')
    buf = BytesIO()
    wb.save(buf)

    with pytest.raises(ValueError, match='[Cc]olumna'):
        parse_centralizer_workbook(buf.getvalue())


def test_parse_centralizer_workbook_rejects_empty_business_name_and_full_name():
    # Regresión: una fila con RUT válido pero 'Nombre / Razón Social' (o
    # 'Nombre Completo') vacío pasaba el parseo y llegaba al INSERT, donde
    # la columna NOT NULL de destino levantaba un 500 sin capturar en vez de
    # un parse_error limpio. Confirma que ahora se descarta a parse_errors,
    # igual que RUT/Patente vacíos, sin bloquear las demás filas válidas.
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()
    ws_empresas = wb.active
    ws_empresas.title = 'Empresas'
    ws_empresas.append(['Nombre / Razón Social', 'RUT', 'DV'])
    ws_empresas.append(['', '99999005', '3'])  # business_name vacío -> parse_error
    ws_empresas.append(['Transportes Válida SPA', '99999006', '4'])  # válida

    ws_conductores = wb.create_sheet('Conductores')
    ws_conductores.append(['RUT Empresa', 'DV Empresa', 'Nombre Completo', 'RUT Conductor', 'DV Conductor'])
    ws_conductores.append(['99999006', '4', '', '99999201', '2'])  # full_name vacío -> parse_error
    ws_conductores.append(['99999006', '4', 'Conductor Válido', '99999202', '3'])  # válido

    wb.create_sheet('Vehiculos_Equipos')

    buf = BytesIO()
    wb.save(buf)

    result = parse_centralizer_workbook(buf.getvalue())

    assert len(result['transporters']) == 1
    assert result['transporters'][0]['business_name'] == 'Transportes Válida SPA'
    assert len(result['drivers']) == 1
    assert result['drivers'][0]['full_name'] == 'Conductor Válido'

    reasons = {(e['sheet'], e['reason']) for e in result['parse_errors']}
    assert ('Empresas', "'business_name' vacío, fila omitida") in reasons
    assert ('Conductores', "'full_name' vacío, fila omitida") in reasons
