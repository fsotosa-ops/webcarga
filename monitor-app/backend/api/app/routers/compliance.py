"""public.compliance_records — PATCH libre de status/expiration_date +
endpoint de archivos (H2.4). El upload es el "camino feliz" real: a
diferencia de la implementación vieja (solo persiste storage_path plano y
nunca transiciona status), este SIEMPRE deja status='APPROVED_MANUAL' y
persiste la evidencia en el JSONB metadata — ver context_carriers.md §4.2.
No existe un proceso de due diligence separado del negocio hoy: quien sube
el archivo ya lo revisó, no queda un estado intermedio "en revisión"
(decisión explícita del usuario 2026-07-18) — PENDING_REVIEW sigue siendo
un valor válido del CHECK constraint (datos legacy), pero nada nuevo lo
setea.
"""
import csv
import io
import json
import unicodedata
from collections import Counter
from datetime import date, datetime
from typing import Literal, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import Response

from ..auth import get_current_user, get_supabase, require_editor
from ..db import get_pool
from ..schemas.carrier import ACTIVE_OPERATIONAL_STATUS

# Los estados que el embudo considera "en juego". ONBOARDING es una empresa
# recién creada sin RUT todavía: pertenece a "Recién creadas · sin documentos",
# no al catálogo del fondo.
FUNNEL_ACTIVE_STATUSES = (ACTIVE_OPERATIONAL_STATUS, "ONBOARDING")
from ..schemas.compliance import (
    ReassignBody,
    ComplianceRecordPatchBody,
    ComplianceSummaryResponse,
    PendingComplianceListResponse,
)
from ..schemas.document_ingest import unclassified_predicate
from ..services.audit import log_change, record_manual_edit
from ..services.plantilla_certificacion import (
    COLUMNAS,
    COLUMNA_LLAVE,
    COLUMNA_TENENCIA,
    COLUMNA_VENCIMIENTO,
    FUENTE_AUDITORIA,
    SQL_APLICAR,
    SQL_AUDITAR,
    TENENCIA_NO,
    TENENCIA_SI,
    sql_filas_plantilla,
)
from ..services.vencimientos import por_vencer_predicate, vencido_predicate
from ..utils.document_storage import (
    content_sha256_of_stored_file, delete_document_version, get_document_history,
    log_document_replacement, resolve_signed_url, upload_document_version,
)

router = APIRouter(prefix="/compliance-records", tags=["compliance"])

_CATEGORY_BY_ENTITY_TYPE = {"CARRIER": "EMPRESA", "DRIVER": "CHOFER", "ASSET": "EQUIPO"}


def _certification_type(requirement_level: str) -> str:
    return "BASICA" if requirement_level == "LEGAL_MANDATORY" else "ADICIONAL"


async def _fetch_record(record_id: str, pool, supabase=None) -> dict:
    row = await pool.fetchrow(
        """
        SELECT cr.id, cr.entity_id, cr.entity_type, cr.requirement_id, req.requirement_code, req.name,
               req.requirement_level, req.requires_file, cr.status, cr.expiration_date, cr.file_url,
               cr.metadata, cr.is_manual_override, cr.created_at, cr.updated_at
        FROM public.compliance_records cr
        JOIN public.compliance_requirements req ON req.id = cr.requirement_id
        WHERE cr.id = $1 AND cr.is_current = true
        """,
        record_id,
    )
    if not row:
        raise HTTPException(404, "Registro de cumplimiento no encontrado")
    record = dict(row)
    # file_url guarda el storage_path crudo (ver upload_compliance_file) — el
    # bucket no es público, hay que firmarlo antes de devolverlo al frontend.
    if supabase is not None:
        record["file_url"] = resolve_signed_url(supabase, record["file_url"])
    return record


# Cada agrupación cambia sólo dos cosas: por qué entidad se agrupa y de dónde
# sale su nombre. El resto de la consulta es idéntico, así que se parametriza en
# vez de escribir tres consultas que después divergen.
# UNA definición de "pendiente", usada por el embudo y por /pending.
#
# Un registro cuenta como pendiente por CUALQUIERA de las dos vías: que su
# estado lo diga, o que su fecha ya haya pasado aunque nadie recalculó el
# estado. Verificado contra producción: los 9 registros vencidos por fecha
# tienen status APPROVED_MANUAL, así que mirando sólo el estado el embudo
# mandaba 8 empresas a "Hay que renovar" mientras el cajón de cada una decía
# "No le falta ningún documento" — se pedía renovar algo que la interfaz se
# negaba a nombrar.
def pendiente_predicate(alias: str = "cr") -> str:
    """Lo que le falta a alguien: no tiene el documento, o el que tiene ya no
    sirve, o esta por dejar de servir.

    OJO: este predicado lo comparten /pending, el embudo (GET /status) y el
    cajon. Ya hubo un bug por moverlos por separado (ver el comentario de
    arriba). Si cambias este predicado, las tres lecturas se mueven JUNTAS.

    La tercera via —"por vencer"— se sumo en la Ronda 129. Antes, renovar no
    tenia superficie en ninguna parte: el predicado exigia expiration_date <
    CURRENT_DATE, o sea YA vencido, asi que un documento que vence en diez
    dias no aparecia ni en el cajon ni en la etapa "Hay que renovar" del
    embudo. Medido al aplicarlo: 5.035 -> 5.038 pendientes. Los 3 que entran
    son una poliza de seguro que vence en tres dias y dos revisiones de
    vehiculo, las tres en APPROVED_MANUAL con fecha futura — invisibles hasta
    el dia en que fuera tarde."""
    return (
        f"({alias}.status IN ('MISSING','EXPIRED') "
        f"OR {vencido_predicate(alias)} "
        f"OR {por_vencer_predicate(alias)})"
    )


_STATUS_GROUPS = {
    "carrier": {
        "entity_type": "CARRIER",
        "table": "public.carriers",
        "name_col": "business_name",
    },
    "driver": {
        "entity_type": "DRIVER",
        "table": "public.drivers",
        "name_col": "full_name",
    },
    "asset": {
        "entity_type": "ASSET",
        "table": "public.assets",
        "name_col": "license_plate",
    },
    # La cuarta agrupación no es por entidad sino por TIPO de documento (D2):
    # responde "qué requisito falta más", que mirando empresa por empresa no se
    # ve. Cruza todas las empresas, así que no tiene empresa propia ni etapa de
    # embudo.
    "requirement": {
        "entity_type": None,
        "table": "public.compliance_requirements",
        "name_col": "name",
    },
}


@router.get("/status")
async def get_certification_status(
    group: Literal["carrier", "driver", "asset", "requirement"] = Query("carrier"),
    scope: Literal["active", "catalog"] = Query("active"),
    carrier_id: str | None = Query(None),
    q: str = Query(""),
    limit: int = Query(200, le=500),
    pool=Depends(get_pool),
    _=Depends(get_current_user),
):
    """Cómo va la certificación, agrupada por empresa, conductor o vehículo.

    Es la misma lista mirada de tres maneras (HU-04). Un conductor o un vehículo
    sin la empresa a la que pertenece no dice nada, así que **la fila siempre
    trae su empresa** — cuando se agrupa por empresa, es ella misma.

    Los documentos sin clasificar sólo se cuentan agrupando por empresa: un
    archivo en la bandeja pertenece a una empresa, no a un conductor.

    Con `carrier_id` se acota a la flota de una empresa. Es lo que usa el panel
    de detalle, y a propósito por la misma consulta: así el "N de M" de un
    conductor es idéntico mirándolo desde la lista o desde su empresa.

    `scope` parte el universo de empresas en dos mitades **disjuntas y
    exhaustivas**: `active` son las operativas más cualquiera con documentos
    esperando, y `catalog` es exactamente su complemento (209 empresas hoy). Se
    piden por separado porque juntas no caben en el `limit`, y el embudo muestra
    el catálogo plegado. Que sean complemento y no dos filtros independientes es
    lo que garantiza que ninguna empresa quede invisible ni contada dos veces.
    """
    cfg = _STATUS_GROUPS[group]

    if group == "carrier":
        # La empresa suma lo suyo y lo de sus conductores y vehículos activos,
        # mismo criterio de atribución que el resto del roster.
        entity_join = """
            LEFT JOIN attributed a ON a.carrier_id = e.id
            LEFT JOIN docs d       ON d.carrier_id = e.id
            LEFT JOIN gestion g    ON g.carrier_id = e.id
            LEFT JOIN viajes v     ON v.carrier_id = e.id
        """
        # `$1` (el estado ACTIVE) se pasa SIEMPRE en esta agrupación, incluso
        # acotada a una empresa: el embudo lo referencia para decidir si la fila
        # cae en "Resto del catálogo". Antes se omitía en ese caso y había que
        # renumerar todo lo demás — de ahí salió el bug de placeholders que
        # cubre test_status_binds_exactly_the_parameters_it_references.
        params: list = [list(FUNNEL_ACTIVE_STATUSES), q, limit]
        p_q, p_limit = "$2", "$3"
        # El complemento se escribe negando la MISMA expresión, no repitiendo
        # el criterio invertido a mano: así no pueden divergir.
        # ONBOARDING entra en el alcance activo. Una empresa creada sin RUT
        # queda en ese estado —lo fija el validador de CarrierCreateBody— y
        # dejarla fuera la mandaba a "Resto del catálogo": plegado, al fondo,
        # detrás de 209 empresas. Es el flujo exacto que el embudo vino a
        # servir: crear una empresa y verla arriba, en "Recién creadas".
        es_activa = "(e.operational_status = ANY($1) OR COALESCE(d.unclassified, 0) > 0)"
        if carrier_id:
            params.append(carrier_id)
            entity_where = "e.id = $4::uuid"
        elif scope == "catalog":
            entity_where = f"NOT {es_activa}"
        else:
            entity_where = es_activa
        carrier_cols = "e.id::text AS carrier_id, e.business_name AS carrier_name, e.operational_status"
        group_by = (
            "e.id, e.business_name, e.operational_status, "
            "d.unclassified, g.management_types, v.trips_30d"
        )
        unclassified = "COALESCE(d.unclassified, 0)::int"
        # Sólo agrupando por empresa hay algo que ordenar por acá.
        orden_cola = "COALESCE(d.unclassified, 0) DESC, "
        # Mismo predicado que la cola de la bandeja, de una sola definición: si
        # acá dice 'UNMATCHED' y allá 'NOT IN (COMMITTED, DISCARDED)', esta
        # pestaña muestra "0 sin clasificar" para una empresa cuya bandeja
        # tiene 12 en cuanto alguien escriba AUTO o SUGGESTED.
        extra_cte = f""",
        docs AS (
            SELECT COALESCE(i.carrier_id, b.carrier_id) AS carrier_id, count(*) AS unclassified
            FROM public.document_ingest_items i
            JOIN public.document_ingest_batches b ON b.id = i.batch_id
            WHERE {unclassified_predicate('i')}
            GROUP BY 1
        ),
        -- Tipo de gestión de la empresa: la flota manda cuando existe (36 de
        -- 39 empresas activas), lo declarado cubre el hueco. La expresión NO
        -- se escribe acá: vive en public.carrier_management_types() (migración
        -- 20260816050000), que es la ÚNICA definición del concepto y la que
        -- también leen las cuatro ramas CARRIER de siembra y la vista previa
        -- del recalcular (app/services/requirement_conditions.py). Antes esta
        -- pantalla tenía su propia copia y la condición nueva leía sólo la
        -- columna declarada: dos definiciones del mismo concepto, una que se
        -- mostraba y otra que se aplicaba (defecto C1). Devuelve CÓDIGOS, no
        -- las etiquetas del catálogo, para que la respuesta hable el mismo
        -- vocabulario que acepta POST /carriers y para que un renombre de
        -- etiqueta —hubo dos en dos días— no cambie el contrato de la API.
        gestion AS (
            SELECT c.id AS carrier_id,
                   public.carrier_management_types(c.id) AS management_types
            FROM public.carriers c
        ),
        -- Actividad reciente. Es una MARCA dentro del grupo, no un criterio de
        -- orden: hoy no hay viajes futuros vinculados (0), así que usarla para
        -- priorizar prometería una anticipación que los datos no tienen.
        -- OJO: app.trips se une por fleet_link_id -> trip_fleet_links.id.
        -- No existe trips.trip_id.
        viajes AS (
            SELECT l.carrier_id, count(*) AS trips_30d
            FROM app.trips t
            JOIN app.trip_fleet_links l ON l.id = t.fleet_link_id
            WHERE l.carrier_id IS NOT NULL
              AND t.planning_date >= CURRENT_DATE - 30
            GROUP BY 1
        )"""
    elif group == "requirement":
        # Se apoya en `attributed`, igual que la agrupación por empresa, para
        # mirar EXACTAMENTE el mismo universo: los pendientes de las empresas
        # activas. Verificado contra la base — 424 CARRIER + 939 DRIVER + 997
        # ASSET = 2.360, el mismo total que devuelve agrupando por empresa. Sin
        # el filtro daría 4.895 y cambiar de pestaña cambiaría el total sin
        # ninguna explicación visible, que es justo lo que el spec §4 prohíbe
        # ("el control de agrupación no crea vistas nuevas").
        entity_join = """
            LEFT JOIN attributed a
                   ON a.requirement_id = e.id
                  AND a.carrier_id IN (
                      SELECT id FROM public.carriers WHERE operational_status = $1
                  )
        """
        entity_where = "a.requirement_id IS NOT NULL"
        params = [ACTIVE_OPERATIONAL_STATUS, q, limit]
        p_q, p_limit = "$2", "$3"
        if carrier_id:
            entity_where += " AND a.carrier_id = $4::uuid"
            params.append(carrier_id)
        # Un requisito cruza todas las empresas: no tiene una sola.
        carrier_cols = (
            "NULL::text AS carrier_id, NULL::text AS carrier_name, "
            "NULL::text AS operational_status"
        )
        group_by = "e.id, e.name"
        unclassified = "0"
        orden_cola = ""
        extra_cte = ""
    else:
        asignacion = "driver_assignments" if group == "driver" else "asset_assignments"
        columna = "driver_id" if group == "driver" else "asset_id"
        entity_join = f"""
            LEFT JOIN records r ON r.entity_type = '{cfg["entity_type"]}' AND r.entity_id = e.id
            LEFT JOIN public.{asignacion} asg
                   ON asg.{columna} = e.id AND asg.status = 'ACTIVE'
            LEFT JOIN public.carriers c ON c.id = asg.carrier_id
        """
        entity_where = "r.entity_id IS NOT NULL"
        carrier_cols = "c.id::text AS carrier_id, c.business_name AS carrier_name, c.operational_status"
        group_by = "e.id, e.{name_col}, c.id, c.business_name, c.operational_status".format(**cfg)
        unclassified = "0"
        # OJO: un literal en ORDER BY lo interpreta Postgres como POSICIÓN
        # ordinal ("ORDER BY 0" es un error), así que acá no va.
        orden_cola = ""
        extra_cte = ""
        params = [q, limit]
        p_q, p_limit = "$1", "$2"
        if carrier_id:
            entity_where += " AND asg.carrier_id = $3::uuid"
            params.append(carrier_id)

    # Por empresa y por requisito el agregado sale de `attributed` —las dos
    # necesitan la atribución a empresa—; por conductor y por vehículo, de las
    # filas de la propia entidad.
    fuente = "a" if group in ("carrier", "requirement") else "r"

    # Un registro cuenta como vencido por CUALQUIERA de las dos vías: que
    # alguien lo haya marcado EXPIRED, o que su fecha ya pasó aunque el estado
    # todavía no se haya recalculado. Mirar sólo el estado subreporta.
    # Vencido y pendiente salen de UNA definición compartida con /pending. Si
    # el embudo cuenta la fecha y /pending no, el embudo manda a renovar
    # documentos que el cajón se niega a nombrar.
    pendiente = pendiente_predicate(fuente)
    # La cuarta copia de la regla de vencimiento, ahora tambien del modulo
    # compartido. Suma `status = 'EXPIRED'` a proposito: el embudo cuenta como
    # vencido lo que alguien marco a mano aunque no tenga fecha.
    #
    # OJO, y queda anotado: la etapa `renovar` la decide ESTE predicado, no
    # `pendiente`. Un documento POR VENCER entra a /pending y al cajon, pero no
    # mueve la etapa — medido al aplicarlo, ninguna de las 2 empresas con algo
    # por vencer cambio de etapa. Ampliar `renovar` a lo proximo a vencer
    # cambiaria el significado de una etiqueta que operaciones ya usa, asi que
    # es decision de negocio y no se toma desde aca.
    vencido = f"({fuente}.status = 'EXPIRED' OR {vencido_predicate(fuente)})"
    cubierto = f"({fuente}.status IS NOT NULL AND NOT {pendiente})"

    if group == "carrier":
        # El embudo decide la etapa en SQL, de UNA definición. Calcularlo en el
        # frontend obligaría a repetir el criterio en el conteo del encabezado y
        # en el orden — que es exactamente como divergen dos superficies del
        # mismo dato.
        funnel_cols = f"""
               count(*) FILTER (WHERE {vencido})                                   AS expired_count,
               g.management_types                                                   AS management_types,
               COALESCE(v.trips_30d, 0)::int                                       AS trips_30d,
               CASE
                   WHEN NOT (e.operational_status = ANY($1)
                             OR COALESCE(d.unclassified, 0) > 0)      THEN 'catalogo'
                   WHEN count(*) FILTER (WHERE {vencido}) > 0          THEN 'renovar'
                   WHEN count({fuente}.status) > 0
                        AND count(*) FILTER (WHERE {cubierto})
                            = count({fuente}.status)                   THEN 'al_dia'
                   WHEN count(*) FILTER (WHERE {cubierto}) = 0         THEN 'sin_documentos'
                   ELSE                                                     'en_proceso'
               END                                                                 AS funnel_group,"""
    else:
        # El embudo es de empresas. Un conductor no tiene etapa de certificación
        # propia, y devolver el campo en null invitaría a dibujarlo igual.
        funnel_cols = ""

    rows = await pool.fetch(
        f"""
        WITH records AS (
            SELECT cr.entity_type, cr.entity_id, cr.status, cr.expiration_date,
                   cr.requirement_id, req.requirement_level
            FROM public.compliance_records cr
            JOIN public.compliance_requirements req ON req.id = cr.requirement_id
            WHERE cr.is_current = true
        ),
        attributed AS (
            SELECT r.status, r.requirement_level, r.expiration_date, r.requirement_id,
                CASE r.entity_type
                    WHEN 'CARRIER' THEN r.entity_id
                    WHEN 'DRIVER'  THEN da.carrier_id
                    WHEN 'ASSET'   THEN aa.carrier_id
                END AS carrier_id
            FROM records r
            LEFT JOIN public.driver_assignments da
                ON r.entity_type = 'DRIVER' AND da.driver_id = r.entity_id AND da.status = 'ACTIVE'
            LEFT JOIN public.asset_assignments aa
                ON r.entity_type = 'ASSET' AND aa.asset_id = r.entity_id AND aa.status = 'ACTIVE'
        ){extra_cte}
        SELECT e.id::text AS entity_id, e.{cfg["name_col"]} AS entity_name,
               {carrier_cols},{funnel_cols}
               count({fuente}.status)                                              AS total_count,
               count(*) FILTER (WHERE {cubierto})                                   AS satisfied_count,
               count(*) FILTER (WHERE {pendiente})                                   AS pending_count,
               count(*) FILTER (WHERE {pendiente}
                                  AND {fuente}.requirement_level = 'LEGAL_MANDATORY') AS pending_mandatory,
               {unclassified}                                                      AS unclassified_count
        FROM {cfg["table"]} e
        {entity_join}
        WHERE {entity_where}
          AND e.{cfg["name_col"]} ILIKE '%' || {p_q}::text || '%'
        GROUP BY {group_by}
        -- Primero donde hay trabajo esperando, después lo más incompleto.
        ORDER BY {orden_cola}pending_count DESC, e.{cfg["name_col"]}
        LIMIT {p_limit}
        """,
        *params,
    )
    result = [dict(r) for r in rows]
    return {
        "total_pending": sum(r["pending_count"] for r in result),
        "total_unclassified": sum(r["unclassified_count"] for r in result),
        "rows": result,
    }


def _estado_de_empresa_a_mostrar(carrier_id: Optional[str]) -> Optional[str]:
    """Que estados de empresa entran en la consulta.

    LA COLA DE TRABAJO NO ES EL EXPEDIENTE. Son dos preguntas distintas y
    hasta ahora compartian un filtro:

    - Sin `carrier_id` es la sabana global: "que hay que hacer hoy". Ahi solo
      entran las empresas ACTIVE, y ese filtro existe por un bug medido (5.4):
      antes traia LEGACY_INACTIVE/INACTIVE/ONBOARDING y eran mas de la mitad
      del volumen mostrado. No se toca.
    - Con `carrier_id` es la ficha de UNA empresa: "que paso con esta". Ahi el
      filtro de estado sobra y ademas MIENTE — devolvia cero sujetos y la
      pantalla lo leia como "nunca se le asignaron requisitos". Hoy son 2
      empresas INACTIVE con 24 registros y 207 LEGACY_INACTIVE con 2.484 que
      el modulo no mostraba.

    En gestion de proveedores dar de baja ARCHIVA, no oculta: el historial de
    cumplimiento es material de auditoria, y al reactivar lo primero que se
    necesita ver es que se vencio durante la baja.
    """
    return None if carrier_id else ACTIVE_OPERATIONAL_STATUS


_PENDING_ROWS_SQL = f"""
WITH pending AS (
    SELECT cr.id, cr.entity_type, cr.entity_id, cr.status, cr.expiration_date,
           -- El HECHO de si hay un archivo, para que la pantalla deje de
           -- deducirlo del estado. `status IN ('MISSING','EXPIRED')` se venia
           -- usando como si significara "no tiene archivo", y significa dos
           -- cosas distintas: un 'EXPIRED' SI tiene archivo —vencio porque
           -- alguien lo subio— y un 'REJECTED' puede no tenerlo. Con la
           -- deduccion, la ficha escondia el documento cargado de todo lo
           -- vencido, que es justo lo que esa pantalla vino a hacer visible.
           cr.file_url IS NOT NULL AS tiene_archivo,
           req.id AS requirement_id,
           req.requirement_code, req.name AS document_name, req.requirement_level,
           req.expiration_policy
    FROM public.compliance_records cr
    JOIN public.compliance_requirements req ON req.id = cr.requirement_id
    WHERE cr.is_current = true
      -- El estado deja de estar incrustado. `falta` es el default y reproduce
      -- exactamente el predicado anterior, para que ningun llamador actual
      -- cambie de comportamiento. `todos` es lo que hace posible la ficha:
      -- ver lo que la empresa TIENE y no solo lo que le falta.
      AND CASE $10::text
            WHEN 'todos'      THEN true
            WHEN 'por_vencer' THEN {por_vencer_predicate('cr')}
            WHEN 'al_dia'     THEN NOT {pendiente_predicate('cr')}
            ELSE {pendiente_predicate('cr')}
          END
),
resolved AS (
    SELECT p.*,
        CASE p.entity_type
            WHEN 'CARRIER' THEN p.entity_id
            WHEN 'DRIVER'  THEN da.carrier_id
            WHEN 'ASSET'   THEN aa.carrier_id
        END AS resolved_carrier_id,
        CASE p.entity_type
            WHEN 'DRIVER' THEN d.full_name
            WHEN 'ASSET'  THEN a.license_plate
            ELSE NULL
        END AS subject_name,
        -- Que ES el vehiculo, no solo su patente. Son dos datos distintos y
        -- ninguno reemplaza al otro: `asset_type` esta SIEMPRE (124 de 124
        -- vehiculos: 87 TRACTOCAMION, 37 RAMPLA) y dice el chasis;
        -- `fleet_service_type_id` es el subtipo de carroceria y solo lo tienen
        -- las ramplas (36 de 37; los 87 tractocamiones lo tienen en NULL).
        -- Por eso viajan los dos y la pantalla dibuja uno o dos badges segun
        -- lo que haya, en vez de elegir un campo y mentir cuando falta.
        --
        -- Nulos para CARRIER y DRIVER por construccion: el LEFT JOIN a
        -- `public.assets` ya solo matchea cuando entity_type = 'ASSET'.
        a.asset_type,
        fst.label      AS fleet_service_type_label,
        fst.bg_color   AS fleet_service_type_bg_color,
        fst.text_color AS fleet_service_type_text_color
    FROM pending p
    LEFT JOIN public.driver_assignments da
        ON p.entity_type = 'DRIVER' AND da.driver_id = p.entity_id AND da.status = 'ACTIVE'
    LEFT JOIN public.asset_assignments aa
        ON p.entity_type = 'ASSET' AND aa.asset_id = p.entity_id AND aa.status = 'ACTIVE'
    LEFT JOIN public.drivers d ON p.entity_type = 'DRIVER' AND d.id = p.entity_id
    LEFT JOIN public.assets a ON p.entity_type = 'ASSET' AND a.id = p.entity_id
    -- Los colores salen del catalogo, no del frontend: `app.status_taxonomies`
    -- ya trae bg_color/text_color por fila y es lo que ya leen las otras
    -- pantallas de flota (assets.py, equipment_closures.py, status_report.py).
    LEFT JOIN app.status_taxonomies fst ON fst.id = a.fleet_service_type_id
),
-- Tipo de Operación (Tractoreo/Equipo Completo) a nivel EMPRESA: no existe
-- como columna propia de carriers — se agrega desde los vehículos activos
-- de la empresa (public.assets.webcarga_operation_type_id). Una empresa
-- con flota mixta aparece con ambos valores, no se fuerza uno solo
-- (confirmado con datos reales, Ronda 85 de AGENTLOG: 36 de 80
-- tractocamiones son "Equipo Completo" mientras el resto de la empresa
-- puede ser "Tractoreo").
carrier_operation_types AS (
    SELECT aa.carrier_id, array_agg(DISTINCT wot.label) AS operation_types
    FROM public.asset_assignments aa
    JOIN public.assets a ON a.id = aa.asset_id
    JOIN app.status_taxonomies wot ON wot.id = a.webcarga_operation_type_id
    WHERE aa.status = 'ACTIVE'
    GROUP BY aa.carrier_id
)
SELECT
    r.id::text, r.entity_type, r.entity_id::text, r.subject_name,
    r.requirement_id, r.requirement_code, r.document_name, r.requirement_level, r.status, r.expiration_date,
    r.expiration_policy,
    -- Se proyecta explicitamente porque el SELECT final enumera columnas: el
    -- `p.*` de `resolved` la trae hasta aca, pero si no se nombra no sale.
    r.tiene_archivo,
    -- Mismo motivo que `tiene_archivo`: sin nombrarlas aca no salen de la CTE.
    r.asset_type,
    r.fleet_service_type_label, r.fleet_service_type_bg_color, r.fleet_service_type_text_color,
    -- Por que esta fila esta pendiente. El orden de las ramas importa: se
    -- pregunta primero por lo vencido, porque `por_vencer` ya lo excluye pero
    -- leerlo al reves invitaria a alguien a "simplificar" el predicado y
    -- volver a mezclarlos.
    --
    -- `status = 'EXPIRED'` va en la MISMA rama que la fecha, igual que en el
    -- embudo (ver `vencido` en get_certification_status): sin eso, un registro
    -- marcado vencido a mano y sin fecha saldria como 'FALTA' en el cajon
    -- mientras el embudo lo cuenta como vencido, que es exactamente el desfase
    -- de dos lecturas que este modulo ya tuvo una vez. Hoy son 0 filas —se
    -- midio—, y se escribe asi para que sigan siendo 0 problemas cuando
    -- aparezca la primera.
    --
    -- 'AL_DIA' (Task 4, ronda de arreglo 1): con `estado='falta'` (el default
    -- de siempre) esta rama nunca se alcanzaba porque TODAS las filas que
    -- llegaban aca ya eran pendientes. Desde que `estado='todos'` existe, una
    -- fila cubierta cae en el `CASE` igual que cualquier otra, y sin esta
    -- rama terminaba en el `ELSE 'FALTA'` de abajo — 'FALTA' pasaba a
    -- significar dos cosas a la vez ("falta de verdad" y "no entro en
    -- ninguna de las anteriores"), la misma clase de bug que este modulo ya
    -- tuvo cinco veces con un valor cargando dos sentidos.
    --
    -- Es la MISMA definicion de "al dia" que ya usa la rama 'al_dia' del
    -- `CASE $10::text` de arriba (`NOT pendiente_predicate`) — no una lista
    -- de estados escrita a mano, que es como el embudo y el cajon ya se
    -- desincronizaron una vez. Va PRIMERO, aunque el orden no cambia el
    -- resultado: `pendiente_predicate` arma sus tres vias (`status IN
    -- ('MISSING','EXPIRED')`, `vencido_predicate`, `por_vencer_predicate`)
    -- con OR, asi que si cualquiera de las tres ramas de abajo fuera
    -- verdadera, `pendiente_predicate` ya seria verdadero y esta rama jamas
    -- se cumpliria — son mutuamente excluyentes por construccion. Ponerla
    -- primera es sobre legibilidad: el `CASE` se lee como la particion que
    -- es, "al dia o no", antes de entrar al detalle de POR QUE no lo esta.
    CASE
        WHEN NOT {pendiente_predicate('r')} THEN 'AL_DIA'
        WHEN r.status = 'EXPIRED' OR {vencido_predicate('r')} THEN 'VENCIDO'
        WHEN {por_vencer_predicate('r')} THEN 'POR_VENCER'
        ELSE 'FALTA'
    END AS urgencia,
    c.id::text AS carrier_id, c.business_name AS carrier_name, c.tax_id AS carrier_tax_id,
    COALESCE(cot.operation_types, ARRAY[]::text[]) AS carrier_operation_types,
    count(*) OVER() AS total_count
FROM resolved r
JOIN public.carriers c ON c.id = r.resolved_carrier_id
LEFT JOIN carrier_operation_types cot ON cot.carrier_id = c.id
-- $8 nulo = sin filtro de estado. NO es "traer todo por defecto": lo
-- pasa nulo UNICAMENTE la consulta acotada a UNA empresa (ver
-- `_estado_de_empresa_a_mostrar`). La sabana global sigue pidiendo
-- ACTIVE, que es el bug 5.4 y la razon por la que este filtro existe.
WHERE ($8::text IS NULL OR c.operational_status = $8)
  AND ($1::uuid IS NULL OR c.id = $1)
  AND ($2::text IS NULL OR r.entity_type = $2)
  AND ($3::text IS NULL OR r.requirement_code = $3)
  AND ($4::text IS NULL OR c.business_name ILIKE '%' || $4 || '%' OR r.subject_name ILIKE '%' || $4 || '%')
  AND ($5::text IS NULL OR $5 = ANY(COALESCE(cot.operation_types, ARRAY[]::text[])))
  -- Un sujeto concreto: lo que le falta a ESTE conductor o a ESTE vehículo.
  -- Sin esto, quien quiera el detalle de una persona tiene que pedir el de la
  -- empresa entera y filtrar del lado del cliente — y esa página corta en 200
  -- (hay empresas con 381 pendientes), así que el filtro del cliente opera
  -- sobre una muestra truncada y no lo dice. Se filtra donde están los datos.
  AND ($9::uuid IS NULL OR r.entity_id = $9)
ORDER BY c.business_name, r.entity_type, r.subject_name
LIMIT $6 OFFSET $7
"""


@router.get("/pending", response_model=PendingComplianceListResponse)
async def list_pending_compliance_records(
    carrier_id: Optional[str] = Query(None),
    category: Optional[Literal["CARRIER", "DRIVER", "ASSET"]] = Query(None),
    requirement_code: Optional[str] = Query(None),
    q: str = Query(""),
    operation_type: Optional[Literal["Tractoreo", "Equipo Completo"]] = Query(None),
    entity_id: Optional[str] = Query(
        None,
        description="Acota a un sujeto concreto (un conductor o un vehículo). "
                    "Se usa junto con `category` para el cajón de una persona.",
    ),
    # 500, no 200 (ronda de arreglo 1, Task 4): el tope subio para que un
    # pedido de hasta 500 filas de un solo golpe no volviera 422 antes de
    # llegar al handler -mismo tope que ya usa /status. Ya no hay un llamador
    # fijo pidiendo justo esa cantidad (comentario viejo, corregido en
    # perf/compresion-y-resumen: decia que "la ficha de empresa pide UNA sola
    # vez con estado='todos'", y desde /summary + Tarea 1 de esa rama eso ya
    # no es cierto -la ficha pide el resumen agregado y el detalle de cada
    # sujeto aparte, con `limit=200`). El tope queda como margen, no como
    # numero que alguien dependa de pedir exacto.
    limit: int = Query(50, le=500),
    offset: int = Query(0, ge=0),
    estado: Literal["falta", "por_vencer", "al_dia", "todos"] = Query(
        "falta",
        description="Qué mostrar. `falta` (default) reproduce el comportamiento "
                    "anterior; `todos` es lo que usa la ficha de empresa.",
    ),
    pool=Depends(get_pool),
    _=Depends(get_current_user),
):
    """Módulo Documentos (sábana) — un `compliance_record` pendiente por
    fila, cruzando toda la flota en vez de navegar empresa por empresa. Ver
    docs/superpowers/plans (plan del módulo). Reusa el mismo criterio de
    atribución DRIVER/ASSET→empresa que `pending-summary`, a nivel de fila.
    Solo empresas activas (bug 5.4): antes traía también LEGACY_INACTIVE/
    INACTIVE/ONBOARDING — confirmado contra datos reales que eran más de la
    mitad del volumen mostrado."""
    rows = await pool.fetch(
        _PENDING_ROWS_SQL,
        carrier_id, category, requirement_code, q or None, operation_type, limit, offset,
        _estado_de_empresa_a_mostrar(carrier_id), entity_id, estado,
    )
    total = rows[0]["total_count"] if rows else 0
    result_rows = [
        {
            "id": r["id"],
            "carrier_id": r["carrier_id"],
            "carrier_name": r["carrier_name"],
            "carrier_tax_id": r["carrier_tax_id"],
            "carrier_operation_types": list(r["carrier_operation_types"]),
            "certification_type": _certification_type(r["requirement_level"]),
            "category": _CATEGORY_BY_ENTITY_TYPE[r["entity_type"]],
            "entity_type": r["entity_type"],
            "entity_id": r["entity_id"],
            "subject_name": r["subject_name"],
            "requirement_id": str(r["requirement_id"]),
            "requirement_code": r["requirement_code"],
            "document_name": r["document_name"],
            "status": r["status"],
            "expiration_date": r["expiration_date"],
            "tiene_archivo": r["tiene_archivo"],
            # Por que esta pendiente y que exige su requisito. Los dos los
            # necesita el renglon de carga: la urgencia para ordenar la
            # atencion, la politica para saber si pedir la fecha ANTES de
            # subir (sin ella el renglon preguntaria siempre, y /file
            # rechazaria con 422 despues de haber subido).
            "urgencia": r["urgencia"],
            "expiration_policy": r["expiration_policy"],
        }
        for r in rows
    ]
    return {"total": total, "rows": result_rows}


# No es un parametro publico: le pide a la CTE TODAS las filas de la
# empresa, no las 500 que /pending tope para su listado global. Hoy la
# empresa mas grande tiene 457; 5000 deja margen sin acercarse a convertir
# esto en un escaneo de la tabla.
#
# A nivel de modulo (no dentro del handler) para que `completo` se pueda
# comparar contra el MISMO numero que se le paso a la CTE, y para que un test
# lo pueda importar en vez de repetirlo a mano (hallazgo 3 de la revision
# final: tres comentarios afirmaban "el resumen nunca viene truncado", y era
# falso -SUMMARY_LIMIT entra como LIMIT DENTRO de la CTE, antes del GROUP BY).
SUMMARY_LIMIT = 5000


# Reusa _PENDING_ROWS_SQL como CTE y agrupa sobre `urgencia`, que ya trae sus
# cuatro ramas resueltas (ver `pendiente_predicate`). El agrupado NO vuelve a
# decidir que es "pendiente" o "al dia" -esa es la misma clase de bug que
# este modulo ya tuvo con el embudo y el cajon contradiciendose.
#
# `falta` agrupa VENCIDO y FALTA: son las dos ramas que la ficha ya muestra
# juntas como "lo que falta" (`avanceDelSujeto`, frontend). Con esta
# particion, al_dia + por_vencer + falta == todos siempre -es lo que el test
# de integracion verifica contra Postgres real.
_SUMMARY_SQL = f"""
WITH pending_rows AS (
{_PENDING_ROWS_SQL}
)
SELECT
    entity_type, entity_id, subject_name, carrier_operation_types,
    asset_type, fleet_service_type_label,
    fleet_service_type_bg_color, fleet_service_type_text_color,
    count(*) AS todos,
    count(*) FILTER (WHERE urgencia = 'AL_DIA') AS al_dia,
    count(*) FILTER (WHERE urgencia = 'POR_VENCER') AS por_vencer,
    count(*) FILTER (WHERE urgencia IN ('FALTA', 'VENCIDO')) AS falta
FROM pending_rows
-- `carrier_operation_types` es constante en toda la respuesta -esta consulta
-- va acotada a UNA empresa (c.id = $1)-, asi que sumarla al agrupado no
-- parte ningun sujeto en dos filas: solo evita una segunda consulta para un
-- dato que ya viaja en cada fila de `_PENDING_ROWS_SQL`.
--
-- Los cuatro campos de vehiculo entran al agrupado por la misma razon y con
-- la misma garantia, pero MAS fuerte: dependen funcionalmente de `entity_id`
-- -son atributos de ESE vehiculo-, asi que no pueden partir un sujeto en dos
-- filas ni aunque la consulta dejara de estar acotada a una empresa. Se
-- agrupan en vez de envolverlos en un `max()` para que eso quede dicho: si
-- alguna vez uno de ellos partiera un sujeto, seria una senal real de que el
-- dato cambio de grano, no algo que un agregado deba tapar.
GROUP BY entity_type, entity_id, subject_name, carrier_operation_types,
         asset_type, fleet_service_type_label,
         fleet_service_type_bg_color, fleet_service_type_text_color
ORDER BY entity_type, subject_name
"""


@router.get("/summary", response_model=ComplianceSummaryResponse)
async def get_compliance_summary(
    carrier_id: str = Query(...),
    pool=Depends(get_pool),
    _=Depends(get_current_user),
):
    """La ficha de una empresa, resumida: cuantos requisitos tiene cada
    sujeto -la empresa, sus conductores, sus vehiculos- y como vienen, sin
    bajar el detalle de cada uno. Reemplaza el fetch de 457 filas que la
    ficha hacia solo para dibujar nueve cabeceras plegadas con sus conteos
    (medido en dev: 57.183 bytes en la primera carga).

    El detalle de un sujeto se pide aparte, solo al desplegarlo
    (GET /pending?entity_id=...) -esta ruta nunca lo trae.

    `completo` (hallazgo 3 de la revision final): SUMMARY_LIMIT entra como
    LIMIT de la CTE, ANTES del GROUP BY -si una empresa superara ese numero
    de registros, el agrupado contaria sobre una lista recortada y las
    cuatro cifras quedarian mal, EN SILENCIO. La forma barata de detectarlo
    sin una segunda consulta: si la suma de `todos` toca el tope exacto, la
    CTE se corto -ninguna empresa real tiene un numero de requisitos que
    coincida justo con SUMMARY_LIMIT salvo que el LIMIT la haya cortado ahi.
    """
    filas = await pool.fetch(
        _SUMMARY_SQL,
        carrier_id, None, None, None, None, SUMMARY_LIMIT, 0,
        _estado_de_empresa_a_mostrar(carrier_id), None, "todos",
    )
    sujetos = [
        {
            "entity_type": f["entity_type"],
            "entity_id": f["entity_id"],
            "subject_name": f["subject_name"],
            "todos": f["todos"],
            "al_dia": f["al_dia"],
            "por_vencer": f["por_vencer"],
            "falta": f["falta"],
            "asset_type": f["asset_type"],
            "fleet_service_type_label": f["fleet_service_type_label"],
            "fleet_service_type_bg_color": f["fleet_service_type_bg_color"],
            "fleet_service_type_text_color": f["fleet_service_type_text_color"],
        }
        for f in filas
    ]
    totales = {
        clave: sum(s[clave] for s in sujetos)
        for clave in ("todos", "al_dia", "por_vencer", "falta")
    }
    return {
        "totales": totales,
        "sujetos": sujetos,
        "completo": totales["todos"] != SUMMARY_LIMIT,
        "carrier_operation_types": list(filas[0]["carrier_operation_types"]) if filas else [],
    }


# ── La planilla de certificación ─────────────────────────────────────────────
#
# OJO CON EL ORDEN: estas rutas van ANTES de `GET /{record_id}`. FastAPI
# resuelve por orden de declaración, así que declaradas después, un GET a
# /date-template entraría por /{record_id} y contestaría "Registro de
# cumplimiento no encontrado" — un 404 que no dice nada de la ruta que falta.

_MAX_FILAS_PLANILLA = 20_000

# Los tres formatos que escribe una persona en Chile. `dd-mm-aaaa` es el que
# baja la planilla; los otros dos entran porque Excel reescribe la columna según
# la configuración regional del computador que la abrió, y rechazar por eso
# sería culpar al usuario de una decisión de Excel.
_FORMATOS_DE_FECHA = ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d")


def _normalizar(valor: str) -> str:
    """Minúsculas y sin tildes. La planilla vuelve de Excel, de Google Sheets y
    de quien la escriba a mano: rechazar un "SI" por la tilde sería culpar a la
    persona del teclado."""
    sin_tildes = unicodedata.normalize("NFKD", valor.strip().lower())
    return "".join(c for c in sin_tildes if not unicodedata.combining(c))


def _parsear_fecha(crudo) -> Optional[date]:
    # openpyxl devuelve datetime cuando la celda tiene formato de fecha, así que
    # el valor no siempre llega como texto.
    if isinstance(crudo, datetime):
        return crudo.date()
    if isinstance(crudo, date):
        return crudo
    for formato in _FORMATOS_DE_FECHA:
        try:
            return datetime.strptime(str(crudo).strip(), formato).date()
        except ValueError:
            continue
    return None


# Un tercer resultado además de sí/no/vacío: "escribió algo que no se
# entiende". Con None querría decir "vacío" y la fila se ignoraría en
# silencio en vez de avisarle a la persona que su "Sii" no se leyó.
NO_SE_ENTIENDE = object()


def _parsear_tenencia(crudo: str):
    valor = _normalizar(crudo)
    if not valor:
        return None
    if valor in TENENCIA_SI:
        return True
    if valor in TENENCIA_NO:
        return False
    return NO_SE_ENTIENDE


def _planilla_a_xlsx(filas: list[dict]) -> bytes:
    """El MISMO módulo escribe la planilla y la lee. Esa es la condición para
    que el ida y vuelta cierre: con el formato decidido en el frontend y el
    parser acá, el día que uno de los dos cambie el separador o el quoting, el
    otro se entera con una fila corrupta y sin error.

    XLSX y no CSV porque Excel lo abre sin diálogo de importación, sin romper
    las tildes y sin reescribir la columna de fechas según la configuración
    regional. Las columnas de contexto van BLOQUEADAS: el `id_registro` es la
    llave que devuelve cada fila a su lugar, y una llave editable por accidente
    es una fila que se aplica al documento equivocado."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Certificación"

    encabezado = Font(bold=True, color="FFFFFF")
    fondo_encabezado = PatternFill("solid", fgColor="141D2B")   # el azul del sidebar
    fondo_editable = PatternFill("solid", fgColor="FFF8E1")     # lo que se completa
    borde = Side(style="thin", color="D0D3DD")

    for columna, definicion in enumerate(COLUMNAS, start=1):
        celda = hoja.cell(row=1, column=columna, value=definicion["label"])
        celda.font, celda.fill = encabezado, fondo_encabezado
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.protection = Protection(locked=True)
        hoja.column_dimensions[get_column_letter(columna)].width = definicion["ancho"]

    for numero, fila in enumerate(filas, start=2):
        for columna, definicion in enumerate(COLUMNAS, start=1):
            celda = hoja.cell(row=numero, column=columna, value=fila[definicion["csv_key"]])
            celda.protection = Protection(locked=not definicion["editable"])
            celda.border = Border(bottom=borde)
            if definicion["editable"]:
                celda.fill = fondo_editable

    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS))}{len(filas) + 1}"
    # Sin contraseña a propósito: no es un control de seguridad, es una baranda
    # para que nadie edite la llave sin querer. Quien necesite desbloquearla,
    # puede.
    hoja.protection.sheet = True
    hoja.protection.autoFilter = False
    hoja.protection.sort = False

    memoria = io.BytesIO()
    libro.save(memoria)
    return memoria.getvalue()


def _leer_planilla(nombre: str, crudo: bytes) -> list[dict]:
    """Acepta XLSX y CSV. Baja XLSX porque es lo cómodo, pero alguien la va a
    pasar por Google Sheets y a devolverla como CSV, y rechazarla ahí sería
    hacerle perder el trabajo por el formato."""
    if nombre.lower().endswith((".xlsx", ".xlsm")):
        try:
            libro = load_workbook(io.BytesIO(crudo), read_only=True, data_only=True)
        except Exception:
            raise HTTPException(422, "No se pudo abrir el archivo de Excel.")
        hoja = libro.active
        renglones = hoja.iter_rows(values_only=True)
        cabecera = [str(c).strip() if c is not None else "" for c in next(renglones, [])]
        etiqueta_a_clave = {c["label"]: c["csv_key"] for c in COLUMNAS}
        claves = [etiqueta_a_clave.get(t, t) for t in cabecera]
        return [dict(zip(claves, valores)) for valores in renglones]

    try:
        texto = crudo.decode("utf-8-sig")
    except UnicodeDecodeError:
        # El archivo volvió en la codificación de Excel para Windows. Es un caso
        # normal, no un error del que la persona tenga la culpa.
        try:
            texto = crudo.decode("latin-1")
        except UnicodeDecodeError:
            raise HTTPException(422, "No se pudo leer el archivo. Guárdalo como XLSX o CSV UTF-8.")
    return list(csv.DictReader(io.StringIO(texto), delimiter=";"))


async def _filas_de_planilla(pool, alcance: str) -> list[dict]:
    filas = await pool.fetch(sql_filas_plantilla(pendiente_predicate("cr")), alcance)
    return [dict(f) for f in filas]


@router.get("/date-template/resumen")
async def resumen_de_planilla(
    alcance: Literal["activas", "todas"] = Query("activas"),
    pool=Depends(get_pool), _=Depends(get_current_user),
):
    """Qué trae la planilla ANTES de bajarla.

    Existe para que el botón no mienta. La exportación que había pedía 200 filas
    sobre 5.026 pendientes y bajaba el 4% sin decirlo — Fabián lo vio en vivo en
    la reunión del 21/08: *"aquí no te bajó todo. Te bajó poquito."*
    """
    filas = await _filas_de_planilla(pool, alcance)
    empresas = {f["empresa"] for f in filas if f["empresa"]}
    por_entidad = Counter(f["entidad"] for f in filas)
    return {
        "alcance":  alcance,
        "filas":    len(filas),
        "empresas": len(empresas),
        # Los dos ejes, para que la pantalla pueda decir que esto no es sólo de
        # fechas: hoy son 1.326 con vencimiento y 1.044 de sola tenencia, y esas
        # 1.044 son TODAS obligatorias.
        "con_vencimiento": sum(1 for f in filas if f["lleva_vencimiento"]),
        "solo_tenencia":   sum(1 for f in filas if not f["lleva_vencimiento"]),
        "por_entidad": dict(por_entidad),
    }


@router.get("/date-template")
async def bajar_planilla(
    alcance: Literal["activas", "todas"] = Query(
        "activas",
        description="'activas' son las empresas operativas, que es lo que se "
                    "pidió cargar. 'todas' suma el histórico y las filas sin empresa.",
    ),
    pool=Depends(get_pool), _=Depends(get_current_user),
):
    contenido = _planilla_a_xlsx(await _filas_de_planilla(pool, alcance))
    return Response(
        content=contenido,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="certificacion.xlsx"'},
    )


@router.post("/date-template")
async def cargar_planilla(
    file: UploadFile = File(...),
    dry_run: bool = Form(True),
    pool=Depends(get_pool), user=Depends(require_editor),
):
    """Devuelve a su lugar lo que la planilla declara.

    `dry_run=true` (el default) no escribe nada y contesta exactamente lo que
    pasaría. Guardar y aplicar son dos actos — el mismo gesto que ya usa la
    configuración de documentos, y por el mismo motivo: sin esa separación, un
    archivo mal armado escribe sobre miles de registros antes de que nadie vea
    un número.

    UNA FILA VACÍA NO SE TOCA, y una planilla PARCIAL es válida: se puede subir
    sólo las filas de conductores y el resto queda intacto. Como consecuencia
    deliberada, vaciar una celda no borra lo que había — vaciar es "no sé", no
    "no vence".
    """
    crudo = await file.read()
    filas = _leer_planilla(file.filename or "", crudo)
    if len(filas) > _MAX_FILAS_PLANILLA:
        raise HTTPException(422, f"Máximo {_MAX_FILAS_PLANILLA} filas por planilla")

    encabezados = set(filas[0].keys()) if filas else set()
    faltantes = {COLUMNA_LLAVE, COLUMNA_TENENCIA, COLUMNA_VENCIMIENTO} - encabezados
    if faltantes:
        raise HTTPException(
            422,
            f"Al archivo le faltan columnas: {', '.join(sorted(faltantes))}. "
            "Usa la planilla que baja el botón de descarga.",
        )

    errores: list[dict] = []
    vacias = 0
    pedidas: dict[str, dict] = {}
    for numero, fila in enumerate(filas, start=2):  # 1 es el encabezado
        registro_id = str(fila.get(COLUMNA_LLAVE) or "").strip()
        crudo_tenencia = str(fila.get(COLUMNA_TENENCIA) or "").strip()
        crudo_fecha = fila.get(COLUMNA_VENCIMIENTO)
        crudo_fecha = "" if crudo_fecha is None else str(crudo_fecha).strip()

        if not crudo_tenencia and not crudo_fecha:
            vacias += 1
            continue
        if not registro_id:
            errores.append({"fila": numero, "error": "La fila tiene datos pero no tiene registro"})
            continue

        tenencia = _parsear_tenencia(crudo_tenencia)
        if tenencia is NO_SE_ENTIENDE:
            errores.append({"fila": numero, "error": f'No se entiende "{crudo_tenencia}" en Documento recibido. Escribe Sí o No.'})
            continue
        fecha = _parsear_fecha(crudo_fecha) if crudo_fecha else None
        if crudo_fecha and fecha is None:
            errores.append({"fila": numero, "error": f'Fecha no reconocida: "{crudo_fecha}"'})
            continue
        if registro_id in pedidas and pedidas[registro_id] != {"tenencia": tenencia, "fecha": fecha}:
            errores.append({"fila": numero, "error": "El mismo registro aparece dos veces con valores distintos"})
            continue
        pedidas[registro_id] = {"tenencia": tenencia, "fecha": fecha}

    # Una sola vuelta a la base para saber qué existe, qué ya vale eso, y qué
    # documento no admite fecha. Sin esto la vista previa contaría como
    # "cambian" filas que la base va a rechazar.
    actuales = await pool.fetch(
        """
        SELECT cr.id::text AS id_registro, cr.entity_type, cr.entity_id::text,
               cr.status, cr.expiration_date, req.has_expiration, req.name AS tipo_documento
        FROM public.compliance_records cr
        JOIN public.compliance_requirements req ON req.id = cr.requirement_id
        WHERE cr.id = ANY($1::uuid[]) AND cr.is_current = true
        """,
        list(pedidas.keys()),
    ) if pedidas else []
    por_id = {r["id_registro"]: r for r in actuales}

    cambios: list[dict] = []
    sin_cambios = 0
    for registro_id, pedido in pedidas.items():
        actual = por_id.get(registro_id)
        if actual is None:
            errores.append({"fila": None, "registro_id": registro_id,
                            "error": "Ese registro no existe o ya no está vigente"})
            continue
        if pedido["fecha"] is not None and not actual["has_expiration"]:
            errores.append({"fila": None, "registro_id": registro_id,
                            "error": f'"{actual["tipo_documento"]}" no lleva fecha de vencimiento'})
            continue

        estado_nuevo = None
        if pedido["tenencia"] is True:
            estado_nuevo = "APPROVED_MANUAL"
        elif pedido["tenencia"] is False:
            estado_nuevo = "MISSING"

        # Marcar recibido un documento que SÍ vence, sin declarar su fecha, es
        # exactamente el defecto que dejó 14 documentos invisibles: quedan
        # aprobados con expiration_date NULL y desaparecen de pendientes para
        # siempre, aunque el papel real venza el mes que viene.
        vence_final = pedido["fecha"] or actual["expiration_date"]
        if estado_nuevo == "APPROVED_MANUAL" and actual["has_expiration"] and vence_final is None:
            errores.append({"fila": None, "registro_id": registro_id,
                            "error": f'"{actual["tipo_documento"]}" necesita su fecha de vencimiento para darse por recibido'})
            continue

        cambia_estado = estado_nuevo is not None and estado_nuevo != actual["status"]
        cambia_fecha = pedido["fecha"] is not None and pedido["fecha"] != actual["expiration_date"]
        if not cambia_estado and not cambia_fecha:
            sin_cambios += 1
            continue
        cambios.append({
            "id": registro_id,
            "estado": estado_nuevo if cambia_estado else None,
            "fecha": pedido["fecha"] if cambia_fecha else None,
            "estado_antes": actual["status"],
            "fecha_antes": actual["expiration_date"],
            "entity_type": actual["entity_type"],
            "entity_id": actual["entity_id"],
        })

    resumen = {
        "cambian": len(cambios),
        "recibidos": sum(1 for c in cambios if c["estado"] == "APPROVED_MANUAL"),
        "fechas": sum(1 for c in cambios if c["fecha"] is not None),
        "sin_cambios": sin_cambios,
        "vacias": vacias,
        "errores": errores[:50],
        "total_errores": len(errores),
        "aplicado": False,
    }
    if dry_run:
        return resumen

    # Con errores no se escribe nada. Es la misma regla que la carga masiva de
    # viajes ("no se importó nada"): una planilla a medio aplicar deja a la
    # persona sin saber qué quedó adentro.
    if errores:
        raise HTTPException(422, {"message": "La planilla tiene errores — no se aplicó nada",
                                  "errors": errores[:50]})
    if not cambios:
        return resumen

    # Un renglón de auditoría por CAMPO cambiado, no por fila: una fila que
    # mueve los dos ejes son dos hechos distintos.
    aud_tipo, aud_id, aud_campo, aud_antes, aud_despues = [], [], [], [], []
    for c in cambios:
        for campo, antes, despues in (
            ("status", c["estado_antes"], c["estado"]),
            ("expiration_date", c["fecha_antes"], c["fecha"]),
        ):
            if despues is None:
                continue
            aud_tipo.append(c["entity_type"])
            aud_id.append(c["entity_id"])
            aud_campo.append(campo)
            aud_antes.append(antes.isoformat() if hasattr(antes, "isoformat") else antes)
            aud_despues.append(despues.isoformat() if hasattr(despues, "isoformat") else despues)

    async with pool.acquire() as conn:
        async with conn.transaction():
            await conn.fetch(
                SQL_APLICAR,
                [c["id"] for c in cambios],      # $1
                [c["estado"] for c in cambios],  # $2
                [c["fecha"] for c in cambios],   # $3
                user["sub"],                     # $4
            )
            await conn.execute(
                SQL_AUDITAR,
                user["sub"], aud_tipo, aud_id, aud_campo, aud_antes, aud_despues,
                FUENTE_AUDITORIA,
            )
    resumen["aplicado"] = True
    return resumen


@router.get("/{record_id}")
async def get_compliance_record(
    record_id: str, pool=Depends(get_pool), supabase=Depends(get_supabase), _=Depends(get_current_user),
):
    return await _fetch_record(record_id, pool, supabase)


@router.post("/{record_id}/reassign")
async def reassign_compliance_document(
    record_id: str,
    body: ReassignBody,
    pool=Depends(get_pool),
    supabase=Depends(get_supabase),
    user=Depends(require_editor),
):
    """Corrige un documento cargado en el lugar equivocado (HU-03).

    Dos operaciones que cubren las cuatro variantes de la HU:
    - con destino → lo mueve a otro requisito, de la misma entidad o de otra;
    - con `to_tray` → lo devuelve a la bandeja de sin clasificar.

    "A otra empresa" se compone: devolver a la bandeja y después mover el item,
    que es la operación que ya existe.

    **El archivo nunca se copia ni se borra**: viaja el mismo `storage_path`.
    Es lo único irrecuperable del sistema, así que reasignar mueve la
    referencia y nada más.
    """
    tiene_destino = bool(body.target_entity_type and body.target_entity_id and body.target_requirement_id)
    if not tiene_destino and not body.to_tray:
        raise HTTPException(422, "Indica un destino o devuélvelo a sin clasificar")
    if tiene_destino and body.to_tray:
        raise HTTPException(422, "O se reasigna a un requisito o se devuelve a la bandeja, no ambos")

    async with pool.acquire() as conn:
        async with conn.transaction():
            origen = await conn.fetchrow(
                "SELECT id::text, entity_type, entity_id::text, status, expiration_date, "
                "file_url, metadata FROM public.compliance_records "
                "WHERE id = $1 AND is_current = true",
                record_id,
            )
            if not origen:
                raise HTTPException(404, "Registro de cumplimiento no encontrado")
            if not origen["file_url"]:
                raise HTTPException(422, "Ese requisito no tiene ningún archivo que reasignar")

            storage_path = origen["file_url"]
            meta = origen["metadata"] or {}
            if isinstance(meta, str):
                meta = json.loads(meta)

            if body.to_tray:
                # Vuelve a la bandeja de la empresa a la que pertenece hoy.
                carrier_id = await conn.fetchval(
                    """
                    SELECT CASE $2::text
                        WHEN 'CARRIER' THEN $1::uuid
                        WHEN 'DRIVER'  THEN (SELECT carrier_id FROM public.driver_assignments
                                             WHERE driver_id = $1::uuid AND status = 'ACTIVE' LIMIT 1)
                        WHEN 'ASSET'   THEN (SELECT carrier_id FROM public.asset_assignments
                                             WHERE asset_id = $1::uuid AND status = 'ACTIVE' LIMIT 1)
                    END
                    """,
                    origen["entity_id"], origen["entity_type"],
                )
                if not carrier_id:
                    raise HTTPException(
                        422,
                        "No se puede devolver a la bandeja: la entidad no tiene una empresa activa asignada",
                    )
                batch_id = await conn.fetchval(
                    """
                    INSERT INTO public.document_ingest_batches
                        (carrier_id, source, status, created_by, total_files, unmatched)
                    VALUES ($1, 'UPLOAD', 'REVIEW', $2, 1, 1)
                    RETURNING id::text
                    """,
                    carrier_id, user["sub"],
                )
                # El hash VIAJA con el item. Sin el, `mismo_contenido` no ve
                # el caso destructivo: un archivo devuelto a la bandeja y su
                # gemelo byte a byte recien subido se listan los dos como "sin
                # colision", y confirmar cualquiera de los dos pisa al otro.
                # Se calcula del blob que ya esta en storage porque es la
                # unica fuente que existe tambien para los registros
                # historicos; si no se puede leer devuelve None, que la senal
                # sabe leer como "no lo se" en vez de como "no hay colision".
                await conn.execute(
                    """
                    INSERT INTO public.document_ingest_items
                        (batch_id, storage_path, file_name, mime_type, size_bytes,
                         match_status, content_sha256)
                    VALUES ($1::uuid, $2, $3, $4, $5, 'UNMATCHED', $6)
                    """,
                    batch_id, storage_path,
                    meta.get("file_name") or storage_path.rsplit("/", 1)[-1],
                    meta.get("mime_type"), meta.get("size_bytes"),
                    content_sha256_of_stored_file(supabase, storage_path),
                )
            else:
                destino = await conn.fetchrow(
                    """
                    SELECT id::text, entity_id::text, entity_type, status, expiration_date
                    FROM public.compliance_records
                    WHERE entity_id = $1 AND requirement_id = $2 AND is_current = true
                    """,
                    body.target_entity_id, body.target_requirement_id,
                )
                if not destino:
                    raise HTTPException(
                        404,
                        "Esa entidad no tiene ese requisito. Verifica la categoría y el tipo de documento.",
                    )
                await _apply_stored_document(
                    conn, destino["id"],
                    storage_path=storage_path,
                    file_name=meta.get("file_name") or storage_path.rsplit("/", 1)[-1],
                    mime_type=meta.get("mime_type"), size_bytes=meta.get("size_bytes"),
                    expiration_date=origen["expiration_date"], actor=user["sub"],
                    entity_type=destino["entity_type"], entity_id=destino["entity_id"],
                    old_status=destino["status"],
                )

            # El origen queda como estaba antes de la carga equivocada.
            await conn.execute(
                """
                UPDATE public.compliance_records SET
                    status = 'MISSING', file_url = NULL, metadata = '{}'::jsonb,
                    expiration_date = NULL, updated_at = NOW()
                WHERE id = $1
                """,
                record_id,
            )
            await log_change(
                conn, actor=user["sub"], entity_type=origen["entity_type"],
                entity_id=origen["entity_id"], action="document_reassign",
                field="file_url", old_value=storage_path,
                new_value="bandeja" if body.to_tray else body.target_requirement_id,
            )

    return {"ok": True, "to_tray": body.to_tray}


@router.patch("/{record_id}")
async def patch_compliance_record(
    record_id: str, body: ComplianceRecordPatchBody, pool=Depends(get_pool),
    supabase=Depends(get_supabase), user=Depends(require_editor),
):
    """Override manual libre (ej. un admin aprueba a mano sin archivo). Para
    subir evidencia real, usar POST /{record_id}/file — ese fuerza
    APPROVED_MANUAL en vez de dejar setear cualquier status a mano."""
    async with pool.acquire() as conn:
        async with conn.transaction():
            current = await conn.fetchrow(
                "SELECT entity_id, entity_type, status, expiration_date FROM public.compliance_records "
                "WHERE id = $1 AND is_current = true",
                record_id,
            )
            if not current:
                raise HTTPException(404, "Registro de cumplimiento no encontrado")

            touched = [f for f in ("status", "expiration_date") if getattr(body, f) is not None]
            if not touched:
                raise HTTPException(422, "Ningún campo enviado")

            await conn.execute(
                """
                UPDATE public.compliance_records SET
                    status = COALESCE($2, status),
                    expiration_date = COALESCE($3, expiration_date),
                    updated_at = NOW()
                WHERE id = $1
                """,
                record_id, body.status, body.expiration_date,
            )
            for field in touched:
                old = current[field]
                new = getattr(body, field)
                await record_manual_edit(
                    conn, table="compliance_records", where={"id": record_id}, actor=user["sub"],
                    entity_type=current["entity_type"], entity_id=current["entity_id"],
                    action="update", field=field,
                    old_value=old.isoformat() if hasattr(old, "isoformat") else old,
                    new_value=new.isoformat() if hasattr(new, "isoformat") else new,
                )
    return await _fetch_record(record_id, pool, supabase)


async def _apply_compliance_upload(
    record_id: str, file: UploadFile, pool, supabase, user,
    expiration_date: date | None = None,
) -> dict:
    """Extraído de upload_compliance_file (endpoint singular) para reusar
    exactamente la misma lógica de status/metadata/auditoría desde el
    endpoint de carga masiva (POST /bulk-file) — un solo lugar que decide
    qué pasa cuando se sube un archivo a un compliance_record.

    `expiration_date` es opcional y se aplica con COALESCE: si no viene, se
    preserva la que ya estuviera declarada. Antes esta función no la escribía
    nunca, así que un documento subido quedaba con la fecha en NULL y — como
    /pending filtra por status — desaparecía de pendientes para siempre."""
    # La politica de vencimiento es del REQUISITO, no del registro, asi que
    # esta consulta —la que ya existia— se extiende con el JOIN en vez de
    # agregar una segunda vuelta a la base.
    current = await pool.fetchrow(
        "SELECT cr.entity_id, cr.entity_type, cr.status, cr.expiration_date, cr.metadata, "
        "       req.expiration_policy "
        "FROM public.compliance_records cr "
        "JOIN public.compliance_requirements req ON req.id = cr.requirement_id "
        "WHERE cr.id = $1 AND cr.is_current = true",
        record_id,
    )
    if not current:
        raise HTTPException(404, "Registro de cumplimiento no encontrado")

    # ANTES de tocar storage. Si validaramos despues, el rechazo dejaria el
    # blob huerfano — que es exactamente el defecto que este trabajo viene a
    # eliminar del camino de la bandeja, donde subir precedia a clasificar y
    # cada 422 dejaba un archivo varado con el requisito vacio.
    if current["expiration_policy"] == "REQUIRED" and expiration_date is None:
        raise HTTPException(422, "Este documento requiere su fecha de vencimiento")

    key_prefix = f"{current['entity_type'].lower()}/{current['entity_id']}/{record_id}"
    uploaded = await upload_document_version(supabase, key_prefix=key_prefix, file=file)

    old_metadata = current["metadata"] or {}
    if isinstance(old_metadata, str):
        old_metadata = json.loads(old_metadata)
    old_storage_path = old_metadata.get("storage_path")
    new_metadata = {
        **old_metadata,
        "storage_path": uploaded["storage_path"],
        "file_name": uploaded["file_name"],
        "mime_type": uploaded["mime_type"],
        "size_bytes": uploaded["size_bytes"],
    }

    async with pool.acquire() as conn:
        async with conn.transaction():
            await conn.execute(
                """
                UPDATE public.compliance_records SET
                    status = 'APPROVED_MANUAL',
                    file_url = $2,
                    metadata = $3::jsonb,
                    expiration_date = COALESCE($4, expiration_date),
                    updated_at = NOW()
                WHERE id = $1
                """,
                record_id, uploaded["storage_path"], json.dumps(new_metadata),
                expiration_date,
            )
            await record_manual_edit(
                conn, table="compliance_records", where={"id": record_id}, actor=user["sub"],
                entity_type=current["entity_type"], entity_id=current["entity_id"],
                action="document_upload", field="status",
                old_value=current["status"], new_value="APPROVED_MANUAL",
            )
            if old_storage_path:
                await log_document_replacement(
                    conn, entity_type=current["entity_type"], entity_id=current["entity_id"],
                    doc_name=f"compliance_record:{record_id}",
                    old_status=current["status"], old_expiry_date=current["expiration_date"],
                    old_storage_path=old_storage_path, actor=user["sub"],
                )

    return {"status": "APPROVED_MANUAL", **uploaded}


async def _apply_stored_document(
    conn, record_id: str, *, storage_path: str, file_name: str,
    mime_type: str | None, size_bytes: int | None,
    expiration_date: date | None, actor: str, entity_type: str, entity_id: str,
    old_status: str,
) -> None:
    """Aplica a un compliance_record un archivo que YA está en storage.

    Contraparte de _apply_compliance_upload para el flujo de la bandeja de sin
    clasificar: ahí el archivo se sube antes de saber a qué requisito
    pertenece, así que cuando se decide ya no hay un UploadFile que leer, solo
    un storage_path. El efecto sobre el record es idéntico (mismo UPDATE,
    misma auditoría) — no se duplica el blob, se referencia el existente.

    Recibe `conn` en vez de `pool` porque el llamador ya está dentro de una
    transacción (marcar el item y aplicar el documento tienen que ser atómicos).

    Si el requisito YA tenía un archivo, registra el reemplazo en audit_log
    antes de pisarlo — igual que `_apply_compliance_upload`. Sin esto el
    puntero al archivo anterior se pierde: el blob sigue en storage (nunca se
    sobrescribe) pero nada permite volver a encontrarlo.
    """
    current = await conn.fetchrow(
        "SELECT metadata, expiration_date FROM public.compliance_records WHERE id = $1",
        record_id,
    )
    old_metadata = (current["metadata"] if current else None) or {}
    if isinstance(old_metadata, str):
        old_metadata = json.loads(old_metadata)
    old_storage_path = old_metadata.get("storage_path")
    if old_storage_path:
        await log_document_replacement(
            conn, entity_type=entity_type, entity_id=entity_id,
            doc_name=f"compliance_record:{record_id}",
            old_status=old_status,
            old_expiry_date=current["expiration_date"] if current else None,
            old_storage_path=old_storage_path, actor=actor,
        )

    metadata = {
        "storage_path": storage_path, "file_name": file_name,
        "mime_type": mime_type, "size_bytes": size_bytes,
    }
    # Deja rastro de que este documento piso a otro. Sin esta marca, deshacer
    # la clasificacion volveria el requisito a MISSING y borraria un documento
    # que era valido antes de la operacion.
    if old_storage_path:
        metadata["replaced_storage_path"] = old_storage_path
    await conn.execute(
        """
        UPDATE public.compliance_records SET
            status = 'APPROVED_MANUAL',
            file_url = $2,
            metadata = $3::jsonb,
            expiration_date = COALESCE($4, expiration_date),
            updated_at = NOW()
        WHERE id = $1
        """,
        record_id, storage_path, json.dumps(metadata), expiration_date,
    )
    await record_manual_edit(
        conn, table="compliance_records", where={"id": record_id}, actor=actor,
        entity_type=entity_type, entity_id=entity_id,
        action="document_upload", field="status",
        old_value=old_status, new_value="APPROVED_MANUAL",
    )


@router.post("/{record_id}/file", status_code=201)
async def upload_compliance_file(
    record_id: str,
    file: UploadFile = File(...),
    expiration_date: Optional[date] = Form(None),
    pool=Depends(get_pool),
    supabase=Depends(get_supabase),
    user=Depends(require_editor),
):
    return await _apply_compliance_upload(record_id, file, pool, supabase, user, expiration_date)


_BULK_UPLOAD_MAX_FILES = 30


@router.post("/bulk-file")
async def bulk_upload_compliance_files(
    carrier_id: str = Form(...),
    record_ids: list[str] = Form(...),
    files: list[UploadFile] = File(...),
    pool=Depends(get_pool),
    supabase=Depends(get_supabase),
    user=Depends(require_editor),
):
    """Módulo Documentos — carga masiva restringida a UNA empresa por vez
    (regla de negocio del diseño validado en Figma: "no es posible subir
    masivamente archivos de varias empresas"). `record_ids`/`files` son dos
    arrays paralelos por índice — el emparejamiento archivo↔record_id lo
    hace el usuario en el modal (arrastra cada archivo a su fila del
    checklist), no hay auto-matching por nombre de archivo (sin precedente
    confiable para eso en esta app). Procesamiento por archivo, NO
    todo-o-nada: un archivo con MIME inválido no tumba el resto del lote —
    mismo criterio que el riesgo documentado en el diseño validado
    ("si el documento se sube en un formato no reconocido, se rechaza y no
    avanza de status", no dice "se rechaza todo el lote")."""
    if len(files) != len(record_ids):
        raise HTTPException(422, "record_ids y files deben tener la misma cantidad de elementos")
    if not files:
        raise HTTPException(422, "No se envió ningún archivo")
    if len(files) > _BULK_UPLOAD_MAX_FILES:
        raise HTTPException(422, f"Máximo {_BULK_UPLOAD_MAX_FILES} archivos por carga masiva")
    if len(set(record_ids)) != len(record_ids):
        raise HTTPException(422, "record_ids duplicados en la misma carga")

    # Defensa en profundidad de "una sola empresa": no confiar solo en que
    # el frontend deshabilite el botón — cada record_id debe resolver al
    # carrier_id recibido (mismo criterio de atribución que /pending).
    owner_rows = await pool.fetch(
        """
        SELECT cr.id::text AS record_id,
            CASE cr.entity_type
                WHEN 'CARRIER' THEN cr.entity_id
                WHEN 'DRIVER'  THEN da.carrier_id
                WHEN 'ASSET'   THEN aa.carrier_id
            END AS resolved_carrier_id
        FROM public.compliance_records cr
        LEFT JOIN public.driver_assignments da
            ON cr.entity_type = 'DRIVER' AND da.driver_id = cr.entity_id AND da.status = 'ACTIVE'
        LEFT JOIN public.asset_assignments aa
            ON cr.entity_type = 'ASSET' AND aa.asset_id = cr.entity_id AND aa.status = 'ACTIVE'
        WHERE cr.id = ANY($1::uuid[])
        """,
        record_ids,
    )
    owner_by_record = {r["record_id"]: r["resolved_carrier_id"] for r in owner_rows}
    mismatched = [
        rid for rid in record_ids
        if str(owner_by_record.get(rid)) != carrier_id
    ]
    if mismatched:
        raise HTTPException(
            422,
            f"Estos registros no pertenecen a la empresa {carrier_id}, no se puede subir en el mismo lote: {mismatched}",
        )

    uploaded, errors = [], []
    for record_id, file in zip(record_ids, files):
        try:
            result = await _apply_compliance_upload(record_id, file, pool, supabase, user)
            uploaded.append({"record_id": record_id, **result})
        except HTTPException as e:
            errors.append({"record_id": record_id, "file_name": file.filename, "error": str(e.detail)})
    return {"uploaded": uploaded, "errors": errors}


@router.delete("/{record_id}/file")
async def delete_compliance_file(
    record_id: str, pool=Depends(get_pool), supabase=Depends(get_supabase), user=Depends(require_editor),
):
    """Borra la evidencia cargada y vuelve el registro a MISSING — mismo
    estado que un documento nunca subido (decisión explícita del usuario
    2026-07-18, no queda un estado "archivado" intermedio)."""
    async with pool.acquire() as conn:
        async with conn.transaction():
            current = await conn.fetchrow(
                "SELECT entity_id, entity_type, status, metadata FROM public.compliance_records "
                "WHERE id = $1 AND is_current = true",
                record_id,
            )
            if not current:
                raise HTTPException(404, "Registro de cumplimiento no encontrado")

            metadata = current["metadata"] or {}
            if isinstance(metadata, str):
                metadata = json.loads(metadata)
            storage_path = metadata.get("storage_path")
            if not storage_path:
                raise HTTPException(422, "Este registro no tiene ningún archivo cargado")

            delete_document_version(supabase, storage_path)

            await conn.execute(
                """
                UPDATE public.compliance_records SET
                    status = 'MISSING',
                    file_url = NULL,
                    metadata = '{}'::jsonb,
                    -- La fecha se va CON el documento. Sin esto el registro
                    -- quedaba MISSING —sin archivo— y con vencimiento futuro:
                    -- un dato que sobrevive a la cosa que describia. La
                    -- `urgencia` de /pending sale de esa fecha, asi que al
                    -- acercarse un documento QUE NO EXISTE aparecia como
                    -- 'POR_VENCER' ("hay que renovarlo") en vez de 'FALTA'.
                    -- La otra ruta al mismo estado (`reassign` con `to_tray`)
                    -- si la limpiaba: eran dos caminos haciendo cosas
                    -- distintas para llegar al mismo lugar.
                    expiration_date = NULL,
                    updated_at = NOW()
                WHERE id = $1
                """,
                record_id,
            )
            await record_manual_edit(
                conn, table="compliance_records", where={"id": record_id}, actor=user["sub"],
                entity_type=current["entity_type"], entity_id=current["entity_id"],
                action="document_delete", field="status",
                old_value=current["status"], new_value="MISSING",
            )
    return await _fetch_record(record_id, pool, supabase)


@router.get("/{record_id}/files")
async def list_compliance_files(
    record_id: str, pool=Depends(get_pool), supabase=Depends(get_supabase), _=Depends(get_current_user),
):
    current = await pool.fetchrow(
        "SELECT entity_id, entity_type, status, expiration_date, file_url, updated_at, overridden_by "
        "FROM public.compliance_records WHERE id = $1",
        record_id,
    )
    if not current:
        raise HTTPException(404, "Registro de cumplimiento no encontrado")
    return await get_document_history(
        pool, supabase, entity_type=current["entity_type"], entity_id=current["entity_id"],
        doc_name=f"compliance_record:{record_id}",
        current_storage_path=current["file_url"],
        current_status=current["status"],
        current_expiry_date=current["expiration_date"],
        current_updated_at=current["updated_at"],
        current_actor=current["overridden_by"],
    )
