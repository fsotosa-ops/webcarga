"""Parser del Excel EETT (Estatus Cumplimiento Gobernanza / "Centralizador") a
estructuras Python normalizadas. Reemplaza el pipeline Mage/dbt congelado que
alimentaba `silver.stg_centralizer_*` (ver AGENTLOG.md y
`.superpowers/sdd/task-1-brief.md`).

Mapeo de columnas ES -> doc_code validado contra `app.compliance_doc_catalog`
real (consultado 2026-07-13) y contra los headers reales del Excel de
producción (leídos con openpyxl, sin comprometer datos de la fila de ejemplo
real a este repo).

Columnas del catálogo de compliance que NO tienen columna en el Excel real —no
es un error de parseo, quedan fuera del alcance de este upload por ahora:
  - transporter: validado_gc, pts_contratista
  - driver:      toma_conoc_plan_emergencia, toma_conoc_pts, capacitacion_epp, f30_1
  - vehicle:     resolucion_sanitaria

Decisión de diseño — status derivado por fecha en Vehiculos_Equipos:
Las columnas "P. Circulación", "Re. Técnica", "Gases Contaminantes" y
"Seguro (SOAP)" tienen en el Excel real la misma forma "fecha + doc" que
"Copia C.I (Vencimiento)"/"Licencia (Vencimiento)" en Conductores (headers
confirmados, ninguna de las 4 trae una columna de status explícita separada).
Por simetría se les aplica la misma regla de status derivado documentada en
el brief para drivers: 'ok' si la fecha es None o futura, 'actualizar' si ya
venció (comparado contra date.today()).

Decisión de diseño — "clients" en el dedupe de Empresas:
El Excel real (headers verificados) NO trae una columna de cliente explícita.
La única señal por-cliente disponible en la hoja son "Avance 80/20"/"Avance
Total" (que en `app.transporter_client_accounts` sí varían por client_name).
Al deduplicar RUTs repetidos se preserva un `clients: list[dict]` con esos dos
valores por cada fila de origen, para que el paso de diff/apply (fuera del
alcance de este task) los resuelva contra el client_name real del upload.

Decisión de diseño — "rut"/"dv"/"rut_dv_valid" en vehicles:
La hoja Vehiculos_Equipos no tiene una columna de RUT propia del vehículo (los
vehículos se identifican por Patente); solo trae "RUT Empresa"/"DV Empresa"
como llave foránea (`transporter_rut`). Por lo tanto las filas de `vehicles`
NO traen `rut`/`dv`/`rut_dv_valid` propios — sí traen `transporter_rut`.
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any, TypedDict

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# RUT helpers
# ---------------------------------------------------------------------------

def normalize_rut(raw: Any) -> str:
    """'12.345.678-9' -> '12345678'. Acepta str o valores numéricos (Excel a
    veces guarda el RUT como int/float si la celda no está formateada como
    texto)."""
    if raw is None:
        return ""
    if isinstance(raw, float) and raw.is_integer():
        raw = int(raw)
    s = str(raw).strip().replace(".", "").replace(" ", "")
    if "-" in s:
        s = s.split("-", 1)[0]
    return s.upper()


def rut_dv(rut_norm: str) -> str:
    """Dígito verificador módulo 11 estándar chileno.

    Verificado a mano contra 3 RUTs reales de dominio público (multiplicando
    de derecha a izquierda por la secuencia 2,3,4,5,6,7 cíclica):
      - SII (Servicio de Impuestos Internos): 60.803.000-K
      - BancoEstado: 97.030.000-7
      - Universidad de Chile: 60.910.000-1
    Los tres calzan con esta implementación.
    """
    digits = [int(c) for c in rut_norm if c.isdigit()]
    total = 0
    factors = (2, 3, 4, 5, 6, 7)
    for i, d in enumerate(reversed(digits)):
        total += d * factors[i % 6]
    remainder = 11 - (total % 11)
    if remainder == 11:
        return "0"
    if remainder == 10:
        return "K"
    return str(remainder)


# ---------------------------------------------------------------------------
# Fechas y status de documentos
# ---------------------------------------------------------------------------

_DATE_FORMATS = ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d")


def parse_centralizer_date(raw: Any) -> date | None:
    """Guard: nunca levanta excepción, retorna None ante cualquier valor no
    parseable (celda vacía, texto libre, etc.)."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


_DOC_STATUS_MAP = {
    "OK": "ok",
    "PENDIENTE": "pendiente",
    "FACTIBLE": "factible",
    "N/A": "n_a",
}


def map_doc_status(raw: Any) -> str | None:
    """'OK'->'ok', 'Pendiente'->'pendiente', 'Factible'->'factible',
    'N/A'->'n_a'. Cualquier otro valor (incluido vacío) -> None."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    return _DOC_STATUS_MAP.get(s.upper())


def _status_from_expiry(expiry: date | None) -> str:
    """'ok' si la fecha es None o futura/hoy, 'actualizar' si ya venció."""
    if expiry is None or expiry >= date.today():
        return "ok"
    return "actualizar"


# ---------------------------------------------------------------------------
# Mapeo de columnas por hoja
# ---------------------------------------------------------------------------
# Cada valor es (tipo_columna, destino):
#   "field"           -> entity[destino] = valor crudo de la celda
#   "rut"             -> normaliza y guarda como identidad (rut) de la entidad
#   "dv"              -> guarda como dv de la identidad de la entidad
#   "transporter_rut" -> normaliza y guarda como entity["transporter_rut"] (FK)
#   "kind"            -> Tipo de Equipo -> entity["kind"] + entity["type_label"]
#   "doc"             -> map_doc_status(valor) -> documents[destino] si no es None
#   "date_doc"        -> destino=(campo_fecha, doc_code): parsea fecha nativa +
#                        status derivado (siempre presente en documents)
#   "ignore"          -> columna reconocida pero descartada (ej. DV Empresa,
#                        que no se usa como llave — la llave es solo el body)

EMPRESAS_COLUMNS: dict[str, tuple[str, Any]] = {
    "Nombre / Razón Social": ("field", "business_name"),
    "RUT": ("rut", None),
    "DV": ("dv", None),
    "ROL SII": ("doc", "rol_sii"),
    "Copia C.I Rep. Legal": ("doc", "copia_ci_rep_legal"),
    "ANEXO RepLeg (GC)": ("doc", "anexo_repleg_gc"),
    "ANEXO 2 (Walmart)": ("doc", "anexo_2_gc"),
    "Contrato WEBCARGA": ("doc", "contrato_webcarga"),
    "F30 (MULTAS)": ("doc", "f30_multas"),
    "F43": ("doc", "f43"),
    "Política de Seguridad": ("doc", "politica_seguridad"),
    "Cert. Afiliación Mutual": ("doc", "cert_mutual"),
    "RIOHS timbrado": ("doc", "riohs_timbrado"),
    "Creación en Walmart": ("doc", "creacion_gc"),
    "Creación en GC": ("doc", "creacion_gc"),  # renombrado en el Excel real (2026-07), mismo doc_code
    "Carpeta Tributaria": ("doc", "carpeta_tributaria"),
    "Cuenta Empresa": ("doc", "cuenta_empresa"),
    "Cuenta Banco Empresa": ("doc", "cuenta_empresa"),  # renombrado en el Excel real (2026-07), mismo doc_code
    "Avance 80/20": ("field", "avance_80_20"),
    "Avance Total": ("field", "avance_total"),
    # Columna opcional — no aparece en el Excel real actual (headers
    # verificados); si en el futuro se agrega, ya queda mapeada y no bloquea
    # el parseo por su ausencia hoy.
    "ID Legado": ("field", "admin_internal_id"),
    "ID Interno Admin": ("field", "admin_internal_id"),
}

CONDUCTORES_COLUMNS: dict[str, tuple[str, Any]] = {
    "RUT Empresa": ("transporter_rut", None),
    "DV Empresa": ("ignore", None),
    "Nombre Completo": ("field", "full_name"),
    "RUT Conductor": ("rut", None),
    "DV Conductor": ("dv", None),
    "Copia C.I (Vencimiento)": ("date_doc", ("id_expiry", "copia_ci")),
    "Licencia (Vencimiento)": ("date_doc", ("license_expiry", "licencia")),
    "ANEXO 3 (Walmart)": ("doc", "anexo_3_gc"),
    "ANEXO GC para Conductor": ("doc", "anexo_3_gc"),  # renombrado en el Excel real (2026-07), mismo doc_code
    "EPP": ("doc", "epp"),
    "DAS / ODI": ("doc", "das_odi"),
    "Hoja de Vida": ("doc", "hoja_de_vida"),
    "Cert. Antecedentes": ("doc", "cert_antecedentes"),
    "Validado por Walmart": ("doc", "validado_gc_driver"),
    "Validado por GC": ("doc", "validado_gc_driver"),  # renombrado en el Excel real (2026-07), mismo doc_code
    "Contrato de Trabajo": ("doc", "contrato_trabajo"),
    "Creación en Walmart": ("doc", "creacion_gc_driver"),
    "Creación en GC": ("doc", "creacion_gc_driver"),  # renombrado en el Excel real (2026-07), mismo doc_code
    "Avance Total": ("field", "avance_total"),
}

VEHICULOS_COLUMNS: dict[str, tuple[str, Any]] = {
    "RUT Empresa": ("transporter_rut", None),
    "DV Empresa": ("ignore", None),
    "Tipo de Equipo": ("kind", None),
    "Patente": ("field", "plate"),
    "Padrón": ("doc", "padron"),
    "P. Circulación": ("date_doc", ("circ_permit_expiry", "permiso_circulacion")),
    "Re. Técnica": ("date_doc", ("tech_inspection_expiry", "revision_tecnica")),
    "Gases Contaminantes": ("date_doc", ("gas_emissions_expiry", "gases")),
    "Seguro (SOAP)": ("date_doc", ("soap_insurance_expiry", "soap")),
    "Póliza Vehicular con RC": ("doc", "poliza_rc"),
    "Año": ("field", "year"),
    "GPS": ("doc", "gps"),
    "Seguro de Carga": ("doc", "seguro_carga"),
    "Mantención Cámara Frío": ("doc", "mantencion_camara_frio"),
    "Creación en Walmart": ("doc", "creacion_gc_vehicle"),
    "Creación en GC": ("doc", "creacion_gc_vehicle"),  # renombrado en el Excel real (2026-07), mismo doc_code
}

_KIND_MAP = {"TRACTOCAMION": "tracto", "RAMPLA": "rampla"}

_SHEET_COLUMNS: dict[str, dict[str, tuple[str, Any]]] = {
    "Empresas": EMPRESAS_COLUMNS,
    "Conductores": CONDUCTORES_COLUMNS,
    "Vehiculos_Equipos": VEHICULOS_COLUMNS,
}


class ParsedUpload(TypedDict):
    transporters: list[dict]
    drivers: list[dict]
    vehicles: list[dict]
    sheet_summary: dict[str, int]
    parse_errors: list[dict]


# ---------------------------------------------------------------------------
# Parseo de una hoja
# ---------------------------------------------------------------------------

def _parse_sheet_rows(
    ws,
    sheet_name: str,
    column_map: dict[str, tuple[str, Any]],
    identity_kind: str,
    required_field: str | None = None,
) -> tuple[list[dict], list[dict]]:
    """identity_kind: 'rut' (Empresas/Conductores) o 'plate' (Vehiculos_Equipos)
    — determina qué campo se exige no-vacío para aceptar la fila.

    required_field: nombre de campo nativo NOT NULL en la tabla destino
    (p.ej. 'business_name'/'full_name') que, si viene vacío, debe rechazar
    la fila hacia parse_errors en vez de dejarla pasar hacia el INSERT — la
    columna destino es NOT NULL, así que dejarla pasar produce un 500 sin
    capturar en vez de un error de parseo limpio."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [c.value for c in header_row]

    for h in headers:
        if h is None or str(h).strip() == "":
            continue
        if h not in column_map:
            raise ValueError(
                f"Columna no mapeada en hoja '{sheet_name}': '{h}'. "
                "Si es una columna nueva legítima, agregarla al mapeo en "
                "centralizer_parser.py; si no, corregir el Excel de origen."
            )

    rows_out: list[dict] = []
    errors: list[dict] = []

    for row_idx, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
        values = {
            headers[i]: row_cells[i].value
            for i in range(min(len(headers), len(row_cells)))
            if headers[i]
        }
        if all(v is None or str(v).strip() == "" for v in values.values()):
            continue  # fila completamente vacía (común al final de la hoja)

        entity: dict[str, Any] = {"documents": {}}
        rut_body = ""
        rut_dv_raw = ""

        for header, raw_value in values.items():
            ctype, target = column_map[header]
            if ctype == "ignore":
                continue
            elif ctype == "field":
                entity[target] = raw_value
            elif ctype == "rut":
                rut_body = normalize_rut(raw_value)
            elif ctype == "dv":
                rut_dv_raw = str(raw_value).strip().upper() if raw_value is not None else ""
            elif ctype == "transporter_rut":
                entity["transporter_rut"] = normalize_rut(raw_value)
            elif ctype == "kind":
                raw_str = str(raw_value).strip() if raw_value is not None else ""
                entity["kind"] = _KIND_MAP.get(raw_str.upper(), "otro")
                entity["type_label"] = raw_str
            elif ctype == "doc":
                status = map_doc_status(raw_value)
                if status is not None:
                    entity["documents"][target] = status
            elif ctype == "date_doc":
                date_field, doc_code = target
                parsed = parse_centralizer_date(raw_value)
                entity[date_field] = parsed
                entity["documents"][doc_code] = _status_from_expiry(parsed)

        if identity_kind == "rut":
            if not rut_body:
                errors.append({
                    "sheet": sheet_name, "row": row_idx,
                    "reason": "RUT vacío o no parseable, fila omitida",
                })
                continue
            entity["rut"] = rut_body
            entity["dv"] = rut_dv_raw
            entity["rut_dv_valid"] = (rut_dv(rut_body) == rut_dv_raw) if rut_dv_raw else False
        else:  # identity_kind == "plate"
            if not entity.get("plate"):
                errors.append({
                    "sheet": sheet_name, "row": row_idx,
                    "reason": "Patente vacía, fila omitida",
                })
                continue

        if required_field and not str(entity.get(required_field) or "").strip():
            errors.append({
                "sheet": sheet_name, "row": row_idx,
                "reason": f"'{required_field}' vacío, fila omitida",
            })
            continue

        entity["_row"] = row_idx
        rows_out.append(entity)

    return rows_out, errors


def _completeness(row: dict) -> int:
    n = sum(1 for k, v in row.items() if k not in ("documents", "_row") and v not in (None, ""))
    n += sum(1 for v in row["documents"].values() if v is not None)
    return n


def _dedupe_transporters(rows: list[dict]) -> list[dict]:
    """Agrupa filas de Empresas repetidas por el mismo RUT (multi-cliente). El
    Excel no trae columna de cliente explícita — ver decisión de diseño en el
    docstring del módulo. La fila con más campos no vacíos se usa como base
    cuando hay conflicto de valores."""
    grouped: dict[str, list[dict]] = {}
    order: list[str] = []
    for row in rows:
        key = row["rut"]
        if key not in grouped:
            grouped[key] = []
            order.append(key)
        grouped[key].append(row)

    result: list[dict] = []
    for key in order:
        group = grouped[key]
        base = max(group, key=_completeness) if len(group) > 1 else group[0]
        merged = dict(base)
        merged["clients"] = [
            {"avance_80_20": g.get("avance_80_20"), "avance_total": g.get("avance_total")}
            for g in group
        ]
        result.append(merged)

    return result


def _get_sheet(wb, name: str):
    if name not in wb.sheetnames:
        raise ValueError(f"Hoja requerida no encontrada en el Excel: '{name}'")
    return wb[name]


def find_unresolved_columns(
    file_bytes: bytes, extra_mappings: dict[str, dict[str, tuple[str, Any]]] | None = None,
) -> list[dict]:
    """Escanea los headers de las 3 hojas contra el mapa combinado (estático
    + extra_mappings, ej. desde app.centralizer_column_mappings) SIN parsear
    filas ni lanzar excepción — usado por el router para decidir si el
    upload puede procesarse directo o necesita la pantalla de mapeo."""
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    extra_mappings = extra_mappings or {}
    unresolved: list[dict] = []
    for sheet_name, column_map in _SHEET_COLUMNS.items():
        merged = {**column_map, **extra_mappings.get(sheet_name, {})}
        ws = _get_sheet(wb, sheet_name)
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        for cell in header_row:
            h = cell.value
            if h is None or str(h).strip() == "":
                continue
            if h not in merged:
                unresolved.append({"sheet": sheet_name, "header": h})
    return unresolved


def parse_centralizer_workbook(
    file_bytes: bytes, extra_mappings: dict[str, dict[str, tuple[str, Any]]] | None = None,
) -> ParsedUpload:
    """Orquesta el parseo de las 3 hojas (Empresas, Conductores,
    Vehiculos_Equipos) del Excel EETT hacia estructuras normalizadas.
    `extra_mappings` (sheet -> {header: (ctype, target)}) se combina con el
    mapa estático de cada hoja — viene de resoluciones guardadas en
    app.centralizer_column_mappings, no requiere tocar este archivo para
    columnas nuevas ya resueltas por un admin."""
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    extra_mappings = extra_mappings or {}

    empresas_rows, empresas_errors = _parse_sheet_rows(
        _get_sheet(wb, "Empresas"), "Empresas",
        {**EMPRESAS_COLUMNS, **extra_mappings.get("Empresas", {})},
        identity_kind="rut", required_field="business_name",
    )
    conductores_rows, conductores_errors = _parse_sheet_rows(
        _get_sheet(wb, "Conductores"), "Conductores",
        {**CONDUCTORES_COLUMNS, **extra_mappings.get("Conductores", {})},
        identity_kind="rut", required_field="full_name",
    )
    vehiculos_rows, vehiculos_errors = _parse_sheet_rows(
        _get_sheet(wb, "Vehiculos_Equipos"), "Vehiculos_Equipos",
        {**VEHICULOS_COLUMNS, **extra_mappings.get("Vehiculos_Equipos", {})},
        identity_kind="plate",
    )

    transporters = _dedupe_transporters(empresas_rows)

    for row in (*transporters, *conductores_rows, *vehiculos_rows):
        row.pop("_row", None)

    return {
        "transporters": transporters,
        "drivers": conductores_rows,
        "vehicles": vehiculos_rows,
        "sheet_summary": {
            "Empresas": len(transporters),
            "Conductores": len(conductores_rows),
            "Vehiculos_Equipos": len(vehiculos_rows),
        },
        "parse_errors": [*empresas_errors, *conductores_errors, *vehiculos_errors],
    }
